"""Generator for the Compensation Management UAT Excel test script.
Extended per phase — append new (id, feature, role, steps, expected) tuples to CASES
as milestones are delivered, then re-run this script to regenerate the workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/Users/elshad/Millers HCM/PRD/HCM_10_Compensation_Management_Multi_Tenant_PRD/Compensation_UAT_Test_Script.xlsx"
FONT = "Arial"
navy = "1F3864"; blue = "2E5496"; band = "F2F5FB"; grey = "808080"; white = "FFFFFF"
thin = Side(style="thin", color="BFBFBF"); border = Border(thin, thin, thin, thin)

# Delivered so far: Phase A (M359 pay bands/config/change-reasons, M360 compensation profile).
# Phases B–F append here as they ship.
CASES = [
 ("SET-01","Setup","HR Admin","Sign in as HR Admin. Make sure at least one salary grade exists with a Min and Max salary (Staffing/Grades), and a test employee 'Test One' who has a current salary and is linked to that grade (via their position).","The grade shows Min/Max; the employee has a salary and a grade."),
 # M359 config
 ("CFG-01","Settings (M359)","HR Admin","Top menu Compensation → Settings. Set 'Max increase % without approval' = 15, 'Budget exceeded policy' = WARNING, 'Default currency' = AZN. Save.","Saved with a green confirmation; reopening the page shows the saved values."),
 # M359 pay bands
 ("PB-01","Pay Bands (M359)","HR Admin","Compensation → Pay Bands → Create. Code=IT-B3, Name=IT Band 3, Grade=(the test grade), Currency=AZN, Min=1500, Mid=2500, Max=3500, Effective From=today. Save.","New row IT-B3 appears with Min 1500 / Mid 2500 / Max 3500 and an Active tag."),
 ("PB-02","Pay Bands (M359)","HR Admin","Create a band with Min=3000, Mid=2000, Max=3500 (min greater than mid). Save.","Rejected — the band range is invalid (min must be ≤ mid ≤ max); an error message is shown, the row is not created."),
 ("PB-03","Pay Bands (M359)","HR Admin","Create another band re-using the code IT-B3. Save.","Rejected — band code must be unique; error shown."),
 ("PB-04","Pay Bands (M359)","HR Admin","On the IT-B3 row click Deactivate and confirm.","The band shows as Inactive (not physically deleted); it no longer counts as an active band."),
 # M359 change reasons
 ("CR-01","Change Reasons (M359)","HR Admin","Compensation → Change Reasons.","Six seeded reasons are present: Annual Merit, Promotion, Market Adjustment, Retention, Equity Adjustment, Correction — each with a category tag."),
 ("CR-02","Change Reasons (M359)","HR Admin","Create a new reason (Code=SPECIAL_ADJ, Name=Special Adjustment, Category=ADJUSTMENT). Save. Then Deactivate it.","The new reason appears, then shows Inactive after deactivation."),
 # M360 profile
 ("CP-01","Compensation Profile (M360)","HR Admin","Compensation → Compensation Profile. In the employee selector search and pick the test employee.","A header card shows the employee name, grade, band and current salary with currency + effective date."),
 ("CP-02","Compensation Profile (M360)","HR Admin","On the profile, read the Compa-ratio card. (Compa-ratio = salary ÷ band midpoint × 100.) Example: salary 2500 and midpoint 2500 → 100%.","The Compa-ratio value is correct for the employee's salary vs the band midpoint (100% when salary equals midpoint)."),
 ("CP-03","Compensation Profile (M360)","HR Admin","Read the Range penetration card. (= (salary − min) ÷ (max − min) × 100.) Example: salary 2500, min 1500, max 3500 → 50%.","The Range-penetration value is correct (50% when the salary sits at the midpoint of the band)."),
 ("CP-04","Compensation Profile (M360)","HR Admin","Look at the range-position tag next to the figures.","The tag colour matches the position: green for MID, gold for LOW/HIGH, red for BELOW_MIN/ABOVE_MAX."),
 ("CP-05","Compensation Profile (M360)","HR Admin","Scroll to the Salary history table.","Past salary records are listed (salary, effective from/to, reason), most recent first — none are missing or overwritten."),
 ("CP-06","Compensation Profile (M360)","Department Manager","Sign in as a Department Manager. Open Compensation → Compensation Profile and try to view an employee who is NOT in your team.","Access is denied for employees outside your hierarchy (an error/permission message); your own team members are viewable."),
 # permissions
 ("SEC-01","Permissions (Phase A)","HR Specialist","Sign in as HR Specialist. Open the Compensation menu.","Can view the Compensation Profile / Pay Bands / Change Reasons, but the create/edit/Save actions and Settings are not available (read-only)."),
 ("SEC-02","Permissions (Phase A)","Employee","Sign in as a plain Employee.","The Compensation admin menu (Pay Bands, Change Reasons, Settings) is not available."),
 # ---------- Phase B: salary-change governance (M361) + exceptions (M362) ----------
 ("SETB","Setup B","HR Admin","Make sure the test employee has a manager (for approval routing) and a pay band covering their grade (from Phase A, e.g. IT-B3 min 1500 / mid 2500 / max 3500).","The employee has a manager and an active pay band.")
 ,("SC-01","Salary Changes (M361)","HR Admin","Compensation → Salary Changes → New salary change. Pick the test employee (their current salary shows), Proposed salary = current + about 5%, Reason = Annual Merit, Effective date = 1st of next month, add a note. Save.","A DRAFT salary-change row appears (Current → Proposed, Increase %, status DRAFT). No red flags because it is within band and under 15%.")
 ,("SC-02","Salary Changes (M361)","HR Admin","On the DRAFT row click Submit.","Status changes to PENDING_APPROVAL and a message says it is now in the Approvals inbox.")
 ,("SC-03","Salary Changes (M361)","HR Admin","Create a new change with Proposed salary ABOVE the band maximum (e.g. 4000 when max is 3500), or an increase over 15%. Save.","The row shows red 'Out-of-band' and/or 'Above-threshold' flags.")
 ,("SC-04","Exceptions (M362)","HR Admin","Go to Compensation → Exceptions.","An OPEN exception was auto-raised for that employee (ABOVE_BAND and/or ABOVE_THRESHOLD) linked to the salary change.")
 ,("SC-05","Approvals (M361)","Manager then HR Admin","Open the Approvals inbox. As the employee's manager approve the pending salary change; then as HR Admin/Compensation Manager approve the second step.","After the final approval the salary-change status becomes APPLIED.")
 ,("SC-06","Salary Changes (M361)","HR Admin","Open the test employee's Compensation Profile (Phase A page).","The profile now shows the NEW salary as current, and the Salary history table has a new row — the OLD salary row is preserved (not overwritten).")
 ,("SC-07","Governance (M361)","HR Admin","Create another salary change but do NOT approve it. Check the employee's Compensation Profile.","Before approval the profile still shows the OLD salary — an unapproved change is never applied / never reaches payroll.")
 ,("SC-08","Self-approval (M361)","Requester","In the Approvals inbox, as the person who raised/owns the change, try to approve your own request.","You cannot approve your own compensation change (self-approval is blocked).")
 ,("SC-09","Salary Changes (M361)","HR Admin","On a DRAFT (not yet submitted) change click Cancel.","Status becomes CANCELLED; nothing is applied.")
 ,("EX-01","Exceptions (M362)","HR Admin","On the Exceptions page, on an OPEN row click Resolve → choose RESOLVED (or WAIVED) and add a note. Save.","The exception status changes to RESOLVED/WAIVED with your note recorded.")
 ,("EX-02","Exceptions (M362)","HR Admin","Use the Status and Exception-type filters on the Exceptions page.","The list filters correctly by status and by exception type.")
 ,("SEC-B1","Permissions (Phase B)","HR Specialist","Sign in as HR Specialist and open Compensation → Salary Changes.","You can view the list but cannot create, submit, or cancel a salary change (read-only).")
 # ---------- Phase C: merit matrix (M363) + budgets (M364) ----------
 ,("MM-01","Merit Matrix (M363)","HR Admin","Compensation → Merit Matrices → open the DEFAULT matrix.","A grid shows 4 performance bands (Excellent/Good/Meets/Below) × 3 range positions (Low/Mid/High) with the seeded percentages: Excellent 8/6/4, Good 5/4/2, Meets 3/2/1, Below 0/0/0.")
 ,("MM-02","Merit Matrix (M363)","HR Admin","Change one cell (e.g. Good/Mid from 4 to 4.5) and Save.","The change is saved and shows the new value when reopened.")
 ,("MM-03","Merit Matrix (M363)","HR Admin","Use the 'Suggest merit' panel: pick the test employee.","It shows the employee's performance band, range position, suggested merit %, current salary, merit amount and new salary. Check: merit amount = current salary × merit % ÷ 100, and new salary = current + merit amount.")
 ,("BUD-01","Budgets (M364)","HR Admin","Compensation → Budgets → Create. Budget Type=MERIT, Scope=GLOBAL, Amount=10000, Currency=AZN, Effective From=today. Save.","The budget appears with Amount 10000, Consumed 0, Remaining 10000.")
 ,("BUD-02","Budgets (M364)","HR Admin","With Settings → Budget policy = WARNING, create a salary change whose increase is larger than the remaining budget.","The salary change is still allowed (WARNING policy only warns, does not block).")
 ,("BUD-03","Budgets (M364)","HR Admin","In Settings set Budget policy = HARD_STOP. Create a salary change whose increase exceeds the remaining MERIT budget.","The salary change is blocked with a 'budget exceeded' error.")
 ,("BUD-04","Budgets (M364)","HR Admin","Set Budget policy back to WARNING. Approve a salary change (via the Approvals inbox) for a budgeted employee, then reopen the MERIT budget.","The budget's Consumed has increased by the salary increase and Remaining has decreased accordingly.")
 ,("BUD-05","Budgets (M364)","HR Admin","On a budget row click Deactivate.","The budget shows Inactive (not physically deleted).")
 # ---------- Phase D: incentives (M365) + commission (M366) ----------
 ,("IN-01","Incentives (M365)","HR Admin","Compensation → Incentives → Plans tab → Create. Code=SALES_INC, Name=Sales Incentive, Measure=SALES, Target %=10, Threshold achievement=80, Target achievement=100, Cap achievement=120, Max payout %=15. Save.","The incentive plan appears in the Plans table.")
 ,("IN-02","Incentives (M365)","HR Admin","Payouts tab → New payout. Pick the plan + test employee, Period=2026-Q1, Achievement %=100. Save.","Payout is created with payout amount = eligible salary × 10% (the target). E.g. salary 2500 → payout 250.")
 ,("IN-03","Incentives (M365)","HR Admin","Create another payout with Achievement %=70 (below the 80 threshold).","Payout amount is 0 (below threshold = no payout).")
 ,("IN-04","Incentives (M365)","HR Admin","Create a payout with Achievement %=120 (at the cap).","Payout uses the maximum payout % (capped); e.g. salary 2500 × 15% = 375.")
 ,("IN-05","Incentives (M365)","HR Admin","On a DRAFT payout click Approve.","Status becomes APPROVED. A note explains it is not attached to payroll until the Payroll-transfer step (Phase E).")
 ,("CM-01","Commission (M366)","HR Admin","Compensation → Commission → Plans → Create. Code=REV_2PCT, Name=Revenue 2%, Basis=REVENUE, Tiered=off, Flat rate %=2. Save.","The commission plan appears.")
 ,("CM-02","Commission (M366)","HR Admin","Payouts tab → New payout. Pick the plan + employee, Period=2026-01, Sales amount=10000. Save.","Commission amount = 200 (2% of 10000).")
 ,("CM-03","Commission (M366)","HR Admin","Create a tiered plan (Tiered=on) with tiers e.g. 0–5000 @1%, 5000+ @3%. Create a payout with Sales amount=10000.","Commission uses marginal tiers: 5000×1% + 5000×3% = 50 + 150 = 200.")
 ,("CM-04","Commission (M366)","HR Admin","On a DRAFT commission payout click Approve.","Status becomes APPROVED (attached to payroll later via the Payroll-transfer step).")
 # ---------- Phase E: market data (M367) + total comp statement (M368) + transfer (M369) ----------
 ,("MK-01","Market Data (M367)","HR Admin","Compensation → Market Data → create a survey (Provider=Mercer, Year=2026). Select it → add a data row (Grade code = the test grade, p25=1800, p50=2400, p75=3000, p90=3600). Then use 'Compare employee' → pick the test employee.","The comparison shows the percentiles, the employee salary, the market ratio (salary ÷ p50 × 100) and a position label (below/at/above market).")
 ,("TC-01","Total Comp (M368)","HR Admin","Compensation → Total Comp Statements → pick the year → Generate all.","A statement per employee appears with base + allowances + bonus + incentives + commission + employer contributions, and a Total that equals the sum of those parts.")
 ,("TC-02","Total Comp (M368)","HR Admin","On a GENERATED statement click Release (or Release all).","Status changes to RELEASED.")
 ,("TC-03","Total Comp (M368)","HR Admin","Click Download on a statement.","A PDF opens showing the compensation breakdown and the total.")
 ,("TC-04","Total Comp (M368)","Employee","Sign in as the Employee → My Workspace → Payroll tab → My total compensation → pick the year → Download.","The employee can download their own RELEASED total-compensation statement only.")
 ,("TR-01","Payroll Transfer (M369)","HR Admin","Compensation → Payroll Transfer. Select an OPEN payroll run (not Paid) and click 'Transfer approved payouts to this run'.","A result shows how many were transferred; the previously-approved incentive/commission payouts become PAID and appear as one-time bonuses on that payroll run; a transfer-log row is recorded.")
 ,("TR-02","Payroll Transfer (M369)","HR Admin","Click Transfer again for the SAME run.","0 payouts are transferred the second time (idempotent — no duplicates).")
 ,("TR-03","Payroll Transfer (M369)","HR Admin","Approve a payout for a TERMINATED employee, then run the transfer.","That payout is skipped (terminated employees are not transferred to payroll).")
 # ---------- Phase F: dashboard/analytics (M370) + letters + notifications (M371) ----------
 ,("DB-01","Dashboard (M370)","HR Admin","Compensation → Dashboard.","Stat cards populate: employees with a comp record, without grade, without band, out of band, pending salary-change approvals, and open exceptions — with sensible numbers.")
 ,("DB-02","Dashboard (M370)","HR Admin","Look at the Compa-ratio distribution chart.","A bar chart shows counts across bands (below-min / low / mid / high / above-max), colour-coded.")
 ,("DB-03","Dashboard (M370)","HR Admin","Look at the Budget utilization table.","Each active budget shows amount, consumed, remaining and a utilization % progress bar.")
 ,("DB-04","Dashboard (M370)","HR Admin","Look at the Out-of-band employees table (also under Reports).","It lists employees whose salary is outside their pay band, with the delta.")
 ,("DB-05","Dashboard (M370)","Department Manager","Sign in as a Department Manager and open the Dashboard.","The figures reflect only your team (hierarchy-scoped), not the whole company.")
 ,("LT-01","Comp Letters (M371)","HR Admin","On Salary Changes, find an APPLIED change and click 'Generate Letter'.","A salary-increase letter PDF opens showing employee name, old salary, new salary, increase % and effective date.")
 ,("LT-02","Comp Letters (M371)","HR Admin","Look at the 'Generate Letter' action on a DRAFT / PENDING change.","It is disabled — a letter is only available once the change is APPLIED.")
 ,("NT-01","Notifications (M371)","Manager / Employee","Submit a salary change (manager is notified), approve it (employee is notified), and trigger an exception (HR/manager notified). Check the notifications bell.","The relevant users receive in-app notifications for submit / approval / exception events.")
]

wb = Workbook()

# Instructions
ws = wb.active; ws.title = "Instructions"; ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 26; ws.column_dimensions["C"].width = 100
def put(r,l,v,bold=True):
    b=ws.cell(r,2,l); b.font=Font(name=FONT,bold=bold,size=11,color=navy if bold else "000000"); b.alignment=Alignment(vertical="top",wrap_text=True)
    c=ws.cell(r,3,v); c.font=Font(name=FONT,size=11); c.alignment=Alignment(vertical="top",wrap_text=True)
ws.cell(1,2,"Compensation Management — User Acceptance Test (UAT) Script").font=Font(name=FONT,bold=True,size=16,color=navy)
ws.cell(2,2,"HCM_10 Compensation Management  •  front-end (browser) testing  •  grows as each phase ships").font=Font(name=FONT,size=11,color=grey)
for r,l,v in [
 (4,"How to use","Open the 'Test Cases' sheet. Do exactly what the Steps say (which menu, which button, what to type), compare to the Expected Result, and pick Pass / Fail / Blocked / Not Run in the Result column. Add anything unusual under Tester Notes. Fill the Sign-off sheet at the end."),
 (6,"Application URL","http://localhost:5180 — sign in first. Compensation features are under the top-nav 'Compensation' menu."),
 (7,"Delivered so far","Phases A–F COMPLETE — Pay Bands, Settings, Change Reasons, Compensation Profile; Salary-change approvals + Exceptions; Merit Matrix + Budgets; Incentives + Commission; Market Data + Total-Comp Statement + Payroll Transfer; Dashboard/Analytics + Comp Letters + Notifications. The whole Compensation module is covered here."),
 (9,"Logins you will need","HR Admin (full compensation access), HR Specialist (read-only), Department Manager (own team only), Employee (self-service). If you only have an admin login, mark the role-restriction rows Blocked with a note."),
 (11,"Result values","Pass = worked as expected.  Fail = did not match (add a note).  Blocked = could not run (missing login/data).  Not Run = skipped."),
 (12,"The two key numbers","Compa-ratio = salary ÷ band midpoint × 100.  Range penetration = (salary − min) ÷ (max − min) × 100. If the employee's salary equals the band midpoint, compa-ratio is 100% and range penetration is 50%."),
]:
    put(r,l,v)
ws.row_dimensions[4].height=45; ws.row_dimensions[7].height=45

# Test Cases
tc = wb.create_sheet("Test Cases"); tc.sheet_view.showGridLines=False
headers=["Test ID","Feature Area","Login As","Test Steps (what to click / type)","Expected Result","Result","Tester Notes"]
widths=[10,22,18,62,58,12,30]
for i,(h,w) in enumerate(zip(headers,widths),1):
    c=tc.cell(1,i,h); c.font=Font(name=FONT,bold=True,color=white,size=11); c.fill=PatternFill("solid",fgColor=navy)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
    tc.column_dimensions[get_column_letter(i)].width=w
tc.freeze_panes="A2"; tc.row_dimensions[1].height=26
for idx,(tid,feat,role,steps,exp) in enumerate(CASES):
    r=idx+2; banded=(idx%2==1)
    for col,v in enumerate([tid,feat,role,steps,exp,"",""],1):
        c=tc.cell(r,col,v)
        c.font=Font(name=FONT,size=11,bold=(col==1),color=(navy if col==1 else "000000"))
        c.alignment=Alignment(vertical="top",wrap_text=True,horizontal=("center" if col in(1,6) else "left"))
        c.border=border
        if banded and col!=6: c.fill=PatternFill("solid",fgColor=band)
    tc.cell(r,6).fill=PatternFill("solid",fgColor="FFF7E6")
    longest=max(len(steps),len(exp)); tc.row_dimensions[r].height=max(30,min(150,(longest//60+1)*15+12))
dv=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True)
dv.add(f"F2:F{len(CASES)+1}"); tc.add_data_validation(dv)

# Sign-off
so=wb.create_sheet("Sign-off"); so.sheet_view.showGridLines=False
for col,w in zip("ABCDE",[3,34,16,22,18]): so.column_dimensions[col].width=w
so.cell(1,2,"Compensation UAT — Sign-off").font=Font(name=FONT,bold=True,size=15,color=navy)
for i,h in enumerate(["Feature Area","Result","Tester","Date"],2):
    c=so.cell(3,i,h); c.font=Font(name=FONT,bold=True,color=white); c.fill=PatternFill("solid",fgColor=blue); c.border=border
areas=["Settings (CFG)","Pay Bands (PB)","Change Reasons (CR)","Compensation Profile (CP)",
       "Salary Changes + Approval (SC)","Compensation Exceptions (EX)",
       "Merit Matrix (MM)","Budgets (BUD)","Incentives (IN)","Commission (CM)","Market Data (MK)","Total Comp (TC)","Payroll Transfer (TR)","Dashboard + Letters (DB/LT/NT)","Permissions (SEC)"]
dv2=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True); so.add_data_validation(dv2)
for j,a in enumerate(areas):
    r=4+j; so.cell(r,2,a).font=Font(name=FONT,size=11)
    for col in range(2,6): so.cell(r,col).border=border
    so.cell(r,3).fill=PatternFill("solid",fgColor="FFF7E6"); dv2.add(so.cell(r,3))
fr=4+len(areas)+2
so.cell(fr,2,"Phase A decision (SHIP / DON'T SHIP):").font=Font(name=FONT,bold=True,size=12,color=navy)
so.cell(fr,3).fill=PatternFill("solid",fgColor="FFF7E6"); so.cell(fr,3).border=border
so.cell(fr+2,2,"Tester name:").font=Font(name=FONT,size=11); so.cell(fr+3,2,"Date:").font=Font(name=FONT,size=11)

wb.save(OUT)
print("WROTE",OUT,"cases:",len(CASES))
