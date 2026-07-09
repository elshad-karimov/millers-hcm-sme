"""Generator for the HCM_29-34 bundle (Assets/Loans/GL/Compliance/Analytics/Engagement)
UAT Excel test script. Append tuples to CASES as milestones ship, re-run to regenerate."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/Users/elshad/Millers HCM/PRD/HCM_29_to_34_Assets_Loans_GL_Compliance_Analytics_Engagement_Multi_Tenant_PRDs/HCM29_34_UAT_Test_Script.xlsx"
FONT = "Arial"
navy = "1F3864"; blue = "2E5496"; band = "F2F5FB"; grey = "808080"; white = "FFFFFF"
thin = Side(style="thin", color="BFBFBF"); border = Border(thin, thin, thin, thin)

CASES = [
 ("SET-01","Setup","HR Admin","Sign in at http://localhost:5180. Check nav: Assets (categories, damage/loss), Payroll (Loan types, Loan requests, GL reconciliation), the new Compliance group, Analytics entries, and Engagement (recognition, pulse schedules, action plans).","All listed entries present."),
 # ---------- 29 Assets ----------
 ("AST-01","Asset categories (M456)","HR Admin","Assets → Categories: create a category (code, name, useful life). Then assign/edit an asset and pick the category.","Category saves and is selectable on the asset form."),
 ("AST-02","Asset transfer (M457)","HR Admin","On an assigned asset use Transfer (to another employee, reason). Approve and complete it in the Transfers tab.","Transfer flows PENDING→APPROVED→COMPLETED; the asset re-assigns and BOTH employees' asset histories show the event."),
 ("AST-03","Damage/loss case (M458)","HR Admin","Assets → Damage/Loss: report a DAMAGE case (asset, employee, est. amount). Approve it with 'propose deduction' ON and an amount.","Case APPROVED and a payroll deduction is created (visible in payroll deductions, type ASSET_DAMAGE). Deduction amount is NOT visible to non-HR roles."),
 ("AST-04","My assets ESS (M459)","Employee","My Workspace → My Assets tab. Use 'Report damage' on one of your assets.","You see only YOUR assets; the report creates a damage case visible to HR; you cannot see other employees' assets."),
 # ---------- 30 Loans ----------
 ("LON-01","Loan types (M460)","HR Admin","Payroll → Loan types: check seeds (GENERAL, EMERGENCY: 3× net cap, 24 months, interest 0%, 6-month tenure, 1 active). Edit a limit.","Seeds present with defaults; edits save (HR Admin only)."),
 ("LON-02","Loan request ESS (M461)","Employee","My Workspace → My Loans → Request a loan (type GENERAL, amount within limits, months, purpose). Then try a second request exceeding 3× your net salary or while one is active.","First request submits (SUBMITTED); the over-limit / second-active request is rejected with a clear eligibility message."),
 ("LON-03","Loan approve → payroll (M461)","HR Admin","Payroll → Loan requests: approve the pending request.","Status APPROVED; a PayrollLoan is created with monthly deduction (check payroll loans/deductions); audited."),
 ("LON-04","Installment schedule (M462)","Employee / HR","As the employee open the loan Statement; as HR expand the request's schedule.","Equal installments with running balance; last installment absorbs rounding; employee sees ONLY their own statement."),
 ("LON-05","Settle / reschedule (M463)","HR Admin","On an approved loan use Settle (partial amount), then Reschedule (new monthly installment).","Outstanding reduces and the remaining schedule regenerates; all steps audited; nothing deleted."),
 ("LON-06","Loan dashboard (M464)","HR Admin","Open the loan dashboard section.","Active loans, total outstanding, completion %, by-department aggregates and overdue list populate. A MANAGER (non-HR) cannot see loan amounts anywhere."),
 # ---------- 31 GL ----------
 ("GL-01","GL approve/post (M465)","HR Admin","Open a payroll run's GL journal: Approve, then Post.","Status DRAFT→APPROVED→POSTED with approvedBy/postedBy stamps; regenerating a POSTED journal is blocked."),
 ("GL-02","GL reversal (M466)","HR Admin","Reverse the POSTED journal.","A new journal with inverted debits/credits appears, linked to the original; original shows REVERSED; reversing twice is blocked."),
 ("GL-03","GL reconciliation (M467)","HR Admin","Payroll → GL reconciliation: pick a period with a posted journal.","Per-component payroll vs GL totals show MATCHED green (or DISCREPANCY red with the difference)."),
 # ---------- 32 Compliance ----------
 ("CMP-01","Statutory templates (M468)","HR Admin","Compliance → Statutory: check the 5 AZ seeds (tax/DSMF/MMI/unemployment monthly due day 20 + annual).","Seeds present; CRUD works."),
 ("CMP-02","Statutory submission (M469)","HR Admin","Create a submission for AZ_TAX_MONTHLY for a period with payroll results → Generate → download the file → mark SUBMITTED then ACCEPTED.","File generates from payroll aggregates and downloads; status transitions DRAFT→GENERATED→SUBMITTED→ACCEPTED are audited."),
 ("CMP-03","Compliance calendar (M470)","HR Admin","Compliance → Calendar: create a deadline; open Upcoming (60 days).","Deadlines list with days-until badges; overdue shows red."),
 ("CMP-04","Work authorization (M471)","HR Admin","Set an employee's work-authorized-until date within 90 days; open Compliance → Work authorization.","The employee appears in the expiring list with days-until (red under 30)."),
 ("CMP-05","Privacy requests (M472)","HR Admin","Create a privacy request (type EXPORT), move OPEN→IN_PROGRESS→COMPLETED with notes.","Transitions work with a +30-day due date; page is HR-Admin-only (HR Specialist/employee cannot open it)."),
 # ---------- 33 Analytics ----------
 ("ANL-01","KPI catalog (M473)","HR Admin","Analytics → KPIs: check ~10 seeds (headcount, turnover, absence, tenure, training, eNPS...). Edit a target value.","Seeds present; edit saves."),
 ("ANL-02","My dashboards (M474)","HR Admin","Analytics → Dashboards: create a layout with 4 KPI widgets and view it.","KPI cards render live values (vs target where set); shared layouts visible to others; unknown codes show n/a."),
 ("ANL-03","Executive summary (M475)","HR Admin","Analytics → Executive.","Cards populate: headcount trend, 12-mo turnover, payroll cost trend, eNPS, upcoming compliance deadlines, attrition high-risk count."),
 ("ANL-04","Attrition risk (M476)","HR Admin / Employee","Analytics → Attrition risk → Recompute. Then sign in as an employee/manager and try the page/API.","Scores 0-100 with factor text (tenure/salary-stagnation/engagement/org-change); HR-Admin-only — everyone else is denied. Anonymous survey data is never used."),
 # ---------- 34 Engagement ----------
 ("ENG-01","Pulse schedules (M477)","HR Admin","Engagement → Pulse schedules: create a WEEKLY schedule for a survey template.","Schedule saves; the daily scheduler creates a campaign when due (verify a campaign appears named per schedule after the run, or trigger date-appropriately)."),
 ("ENG-02","Recognition (M478)","Employee","Engagement → Recognition (or the My Workspace kudos card): give PUBLIC recognition to a colleague (value tag + message). Try recognising yourself.","The kudos appears on the public wall with names and tag colour; self-recognition is rejected; sender is always the logged-in employee (cannot spoof)."),
 ("ENG-03","Recognition moderation (M478)","HR Admin","Hide a wall item as HR, reload the wall.","Hidden item disappears from the wall (audited); unhide restores it."),
 ("ENG-04","Action plans (M479)","HR Admin","Engagement → Action plans: create a plan with 3 items; toggle one done.","Progress shows 1/3; completing all items allows COMPLETED status."),
 ("ENG-05","Participation + anonymity (M480)","HR Admin","Open participation for a campaign where one department has fewer than 5 responses.","Overall rate + by-department rows show; the small department row shows SUPPRESSED with no counts (anonymity guard). Sentiment shows aggregate positive/neutral/negative counts only."),
 # ---------- Cross-cutting ----------
 ("SEC-01","Confidentiality sweep","Employee / Manager","(a) Employee: try another employee's loan statement/assets by id, the attrition-risk page, privacy requests, GL pages. (b) Manager: look for loan amounts or deduction amounts on any team surface.","All denied/absent: loans+GL+deductions+attrition+privacy are HR/Finance-only; employees reach only their own data."),
 ("SEC-02","Payroll boundary","HR Admin","Confirm loan approval/settlement and damage deductions only create rows the payroll run later consumes (no payroll run is executed/changed by these actions).","No payroll run status changes from any bundle action; deductions appear as pending instructions for the next run."),
]

wb = Workbook()
ws = wb.active; ws.title = "Instructions"; ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 26; ws.column_dimensions["C"].width = 100
def put(r,l,v):
    b=ws.cell(r,2,l); b.font=Font(name=FONT,bold=True,size=11,color=navy); b.alignment=Alignment(vertical="top",wrap_text=True)
    c=ws.cell(r,3,v); c.font=Font(name=FONT,size=11); c.alignment=Alignment(vertical="top",wrap_text=True)
ws.cell(1,2,"Assets / Loans / GL / Compliance / Analytics / Engagement (HCM_29-34) — UAT Script").font=Font(name=FONT,bold=True,size=16,color=navy)
ws.cell(2,2,"HCM_29-34 bundle  •  front-end (browser) testing").font=Font(name=FONT,size=11,color=grey)
for r,l,v in [
 (4,"How to use","Open 'Test Cases'. Follow Steps exactly, compare to Expected Result, pick Pass / Fail / Blocked / Not Run. Note oddities in Tester Notes. Fill the Sign-off sheet."),
 (6,"Application URL","http://localhost:5180 — sign in first."),
 (7,"Delivered","29 Assets: categories (M456), transfers (M457), damage/loss + payroll-deduction bridge (M458), My-assets ESS (M459). 30 Loans: type catalog (M460), ESS requests + eligibility + approval→PayrollLoan (M461), installment schedules (M462), settle/reschedule (M463), dashboard (M464). 31 GL: approve/post (M465), reversal (M466), reconciliation (M467). 32 Compliance: AZ statutory templates (M468) + submissions (M469), calendar (M470), work-authorization tracker (M471), privacy requests (M472). 33 Analytics: KPI catalog (M473), saved dashboards (M474), executive summary (M475), attrition risk (M476). 34 Engagement: pulse schedules (M477), recognition wall (M478), action plans (M479), participation + anonymity guard (M480)."),
 (9,"Logins you will need","HR Admin, Manager (confidentiality checks), Employee (ESS + recognition). Mark role-restricted rows Blocked if a login is missing."),
 (11,"Result values","Pass / Fail (note!) / Blocked / Not Run."),
 (12,"KEY RULES under test","(1) Loan amounts + GL data + deduction amounts = HR/Finance-only. (2) Attrition risk = HR-Admin-only, never derived from anonymous surveys. (3) Survey group breakdowns suppressed under 5 responses. (4) Nothing in this bundle runs or changes a payroll run — it only creates instructions the next run consumes. Any breach = automatic Fail + report immediately."),
]:
    put(r,l,v)
ws.row_dimensions[7].height=90; ws.row_dimensions[12].height=60

tc = wb.create_sheet("Test Cases"); tc.sheet_view.showGridLines=False
headers=["Test ID","Feature Area","Login As","Test Steps (what to click / type)","Expected Result","Result","Tester Notes"]
widths=[12,26,20,62,58,12,30]
for i,(h,w) in enumerate(zip(headers,widths),1):
    c=tc.cell(1,i,h); c.font=Font(name=FONT,bold=True,color=white,size=11); c.fill=PatternFill("solid",fgColor=navy)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
    tc.column_dimensions[get_column_letter(i)].width=w
tc.freeze_panes="A2"; tc.row_dimensions[1].height=26
for idx,(tid,feat,role,steps,exp) in enumerate(CASES):
    r=idx+2; banded=(idx%2==1)
    for col,v in enumerate([tid,feat,role,steps,exp,"",""],1):
        c=tc.cell(r,col,v)
        c.font=Font(name=FONT,size=11,bold=(col==1),color=(navy if col==1 else "000000"))
        c.alignment=Alignment(vertical="top",wrap_text=True,horizontal=("center" if col in(1,6) else "left"))
        c.border=border
        if banded and col!=6: c.fill=PatternFill("solid",fgColor=band)
    tc.cell(r,6).fill=PatternFill("solid",fgColor="FFF7E6")
    longest=max(len(steps),len(exp)); tc.row_dimensions[r].height=max(30,min(160,(longest//60+1)*15+12))
dv=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True)
dv.add(f"F2:F{len(CASES)+1}"); tc.add_data_validation(dv)

so=wb.create_sheet("Sign-off"); so.sheet_view.showGridLines=False
for col,w in zip("ABCDE",[3,44,16,22,18]): so.column_dimensions[col].width=w
so.cell(1,2,"HCM_29-34 bundle UAT — Sign-off").font=Font(name=FONT,bold=True,size=15,color=navy)
for i,h in enumerate(["Feature Area","Result","Tester","Date"],2):
    c=so.cell(3,i,h); c.font=Font(name=FONT,bold=True,color=white); c.fill=PatternFill("solid",fgColor=blue); c.border=border
areas=["Asset categories + transfers (AST-01/02)","Damage/loss + deduction bridge (AST-03)","My assets ESS (AST-04)",
       "Loan types (LON-01)","Loan requests + eligibility (LON-02/03)","Installments + settle/reschedule (LON-04/05)",
       "Loan dashboard (LON-06)","GL approve/post/reverse (GL-01/02)","GL reconciliation (GL-03)",
       "Statutory reports (CMP-01/02)","Compliance calendar (CMP-03)","Work authorization (CMP-04)","Privacy requests (CMP-05)",
       "KPI catalog + dashboards (ANL-01/02)","Executive summary (ANL-03)","Attrition risk (ANL-04)",
       "Pulse schedules (ENG-01)","Recognition (ENG-02/03)","Action plans (ENG-04)","Participation + anonymity (ENG-05)",
       "Confidentiality + payroll boundary (SEC)"]
dv2=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True); so.add_data_validation(dv2)
for j,a in enumerate(areas):
    r=4+j; so.cell(r,2,a).font=Font(name=FONT,size=11)
    for col in range(2,6): so.cell(r,col).border=border
    so.cell(r,3).fill=PatternFill("solid",fgColor="FFF7E6"); dv2.add(so.cell(r,3))
fr=4+len(areas)+2
so.cell(fr,2,"Overall decision (SHIP / DON'T SHIP):").font=Font(name=FONT,bold=True,size=12,color=navy)
so.cell(fr,3).fill=PatternFill("solid",fgColor="FFF7E6"); so.cell(fr,3).border=border
so.cell(fr+2,2,"Tester name:").font=Font(name=FONT,size=11); so.cell(fr+3,2,"Date:").font=Font(name=FONT,size=11)

wb.save(OUT)
print("WROTE",OUT,"cases:",len(CASES))
