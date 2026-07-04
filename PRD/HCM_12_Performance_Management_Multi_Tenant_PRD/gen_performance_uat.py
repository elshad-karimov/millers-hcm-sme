"""Generator for the Performance Management UAT Excel test script.
Extended per phase — append new (id, feature, role, steps, expected) tuples to CASES
as milestones are delivered, then re-run this script to regenerate the workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/Users/elshad/Millers HCM/PRD/HCM_12_Performance_Management_Multi_Tenant_PRD/Performance_UAT_Test_Script.xlsx"
FONT = "Arial"
navy = "1F3864"; blue = "2E5496"; band = "F2F5FB"; grey = "808080"; white = "FFFFFF"
thin = Side(style="thin", color="BFBFBF"); border = Border(thin, thin, thin, thin)

# Delivered so far: Phase A — M388 rating scales, M389 review templates + cycle
# scoping, M390 KPI library + assignments, M391 OKR. Phases B–F append here.
CASES = [
 ("SET-01","Setup","HR Admin","Sign in as HR Admin at http://localhost:5180. Open the top-nav 'Performance' menu. You should see (among others): Rating scales, Review templates, KPIs, OKRs, Review cycles, Goals.","The Performance menu shows the new entries: Rating scales, Review templates, KPIs, OKRs."),
 # ---------- Phase A.1: rating scales (M388) ----------
 ("SCL-01","Rating scales (M388)","HR Admin","Performance → Rating scales.","The seeded 'Default 5-point scale' (DEFAULT_5PT) is listed, marked Default + Active, with 5 values: 1 Unsatisfactory, 2 Needs Improvement, 3 Meets Expectations, 4 Exceeds Expectations, 5 Outstanding."),
 ("SCL-02","Rating scales (M388)","HR Admin","Open the default scale (Edit) and look at each value's score band (min–max %).","Bands are ascending and non-overlapping: 0–39.99, 40–59.99, 60–74.99, 75–89.99, 90–100. (These drive the numeric-score → rating conversion in weighted scoring.)"),
 ("SCL-03","Rating scales (M388)","HR Admin","Click 'New scale…'. Code=SIMPLE_3PT, Name=Simple 3-point, Type=Numeric, add 3 values (1 Low / 2 OK / 3 High) with bands 0–49.99 / 50–79.99 / 80–100. Save.","The new scale appears in the list, Active, showing 3 values."),
 ("SCL-04","Rating scales (M388)","HR Admin","Edit SIMPLE_3PT and set overlapping bands (e.g. value 2 = 40–80 while value 1 = 0–49.99). Save.","Rejected — score bands must be ascending and non-overlapping; an error is shown."),
 ("SCL-05","Rating scales (M388)","HR Admin","Edit SIMPLE_3PT and tick 'Default scale'. Save. Reopen the list.","SIMPLE_3PT becomes Default and DEFAULT_5PT loses its Default tag (only one default at a time). Set DEFAULT_5PT back to default afterwards."),
 # ---------- Phase A.2: review templates + cycle scoping (M389) ----------
 ("TPL-01","Review templates (M389)","HR Admin","Performance → Review templates.","The seeded 'Default annual review' (DEFAULT_ANNUAL) template is listed with 8 sections; the scoring-weight total shows 100 (Goals 50 + KPIs 25 + Competencies 20 + Company Values 5; comment/development/final sections carry no weight)."),
 ("TPL-02","Review templates (M389)","HR Admin","Click 'New template…'. Code=SALES_ANNUAL, Name=Sales annual review. Add sections: Goals 40, KPIs 40, Competencies 20, plus Manager Comments (0). Watch the weight tag as you type.","The live scoring-weight tag turns from red to green exactly when the scoring weights total 100; Save succeeds."),
 ("TPL-03","Review templates (M389)","HR Admin","Create another template whose scoring sections total 90 (e.g. Goals 50 + KPIs 40). Save.","Rejected — scoring-section weights must total exactly 100; an error is shown."),
 ("TPL-04","Cycle scoping (M389)","HR Admin","Performance → Review cycles → New cycle (or edit an existing one). Note the new fields: Rating scale, Review template, and audience filters (legal entity / department / grade / employee type / min service months).","The cycle form offers the rating scale and template dropdowns (defaults pre-selected) and the audience-scoping fields; saving keeps them."),
 # ---------- Phase A.3: KPI library + assignments (M390) ----------
 ("KPI-01","KPIs (M390)","HR Admin","Performance → KPIs → Library tab → 'New KPI'. Code=SALES_TGT, Name=Quarterly sales target, Unit=AZN, Default target=50000, Frequency=Quarterly, Scoring=Linear, Source=Manual. Save.","The KPI appears in the library with the LINEAR tag."),
 ("KPI-02","KPIs (M390)","HR Admin","Create a second KPI. Code=CSAT, Name=Customer satisfaction, Unit=%, Default target=90, Scoring=Threshold bands. Save. Then try creating a third KPI re-using code SALES_TGT.","CSAT saves with the THRESHOLD tag; the duplicate SALES_TGT is rejected (code must be unique)."),
 ("KPI-03","KPI assignment (M390)","HR Admin","KPIs → Assignments tab → pick a Review cycle → 'Assign KPI'. KPI=SALES_TGT (target pre-fills 50000), pick an employee, Weight=60. Assign.","The assignment row appears: status ASSIGNED, target 50000, no actual/rating yet."),
 ("KPI-04","KPI measure — linear (M390)","HR Admin","On the SALES_TGT assignment click 'Measure'. Actual value=50000, Period=2026-Q2. Record.","Achievement = 100%, Rating = 5.00 (linear: achievement ÷ 20). Status becomes MEASURED."),
 ("KPI-05","KPI measure — linear partial (M390)","HR Admin","Measure the same assignment again with Actual=30000 (Period=2026-Q3).","Achievement = 60%, Rating = 3.00. Click 'History' — BOTH measurements are listed (Q2 100%/5.00 and Q3 60%/3.00) with who/when."),
 ("KPI-06","KPI measure — threshold (M390)","HR Admin","Assign CSAT (target 90) to an employee, then Measure with Actual=90.","Achievement = 100%, Rating = 4.00 (threshold bands: ≥110→5, ≥100→4, ≥80→3, ≥60→2, else 1 — hitting target = 4; beating it by 10% = 5)."),
 ("KPI-07","KPI duplicate guard (M390)","HR Admin","Try to assign SALES_TGT to the SAME employee in the SAME cycle again.","Rejected — the KPI is already assigned to that employee for this cycle."),
 ("KPI-08","KPI cancel (M390)","HR Admin","Assign any KPI to another employee, then click Cancel on that row and confirm.","Status becomes CANCELLED; Measure is no longer offered on that row."),
 # ---------- Phase A.4: OKR (M391) ----------
 ("OKR-01","OKR objectives (M391)","HR Admin","Performance → OKRs → 'New objective'. Title='Increase market share', Level=Company, pick a cycle, Confidence=High. Save.","The company objective appears with the Company level tag, 0% progress and HIGH confidence."),
 ("OKR-02","OKR alignment (M391)","HR Admin","Create a second objective: Title='Grow enterprise sales', Level=Department, 'Aligns to' = Increase market share. Save.","The department objective appears showing '↑ Increase market share' under its title (parent alignment §8.3)."),
 ("OKR-03","OKR individual owner (M391)","HR Admin","Create a third objective with Level=Individual but leave Owner empty. Save.","Rejected — an Individual objective requires an owner employee. Pick an owner and save again — it appears with the owner's name."),
 ("OKR-04","Key results (M391)","HR Admin","On 'Grow enterprise sales' click '+ KR'. Title='Sign 10 enterprise deals', Measurement=Number, Baseline=0, Target=10, Weight=60. Save. Add a second KR: 'Raise win rate to 40%', Percent, Baseline=20, Target=40, Weight=40.","Expanding the objective row shows both KRs at 0% progress with their weights."),
 ("OKR-05","Check-in + roll-up (M391)","HR Admin","On the objective click 'Check-in'. Key result='Sign 10 enterprise deals', New current value=5, Confidence=Medium, Comment='Half way'. Record.","The KR goes to 50% progress; the OBJECTIVE progress becomes 30% (weighted: 50%×60 + 0%×40 = 30). The success message shows the new progress."),
 ("OKR-06","Second check-in (M391)","HR Admin","Check-in again: KR='Raise win rate to 40%', New current value=30.","That KR shows 50% ((30−20)/(40−20)); objective progress becomes 50% (50×60 + 50×40 = 50). Click History — both check-ins are listed with old → new values, confidence and comments."),
 ("OKR-07","Objective-level comment (M391)","HR Admin","Check-in with NO key result selected — just Confidence=Low and a comment 'Market headwinds'. Record.","The check-in is recorded in History as a comment; the objective's confidence tag changes to LOW; progress is unchanged."),
 ("OKR-08","Close / reopen (M391)","HR Admin","On the objective click Close and confirm. Then try Check-in / Edit.","Status becomes CLOSED; Check-in / + KR / Edit disappear. Reopen restores it to ACTIVE."),
 # ---------- Phase A permissions ----------
 ("SEC-A1","Permissions (Phase A)","HR Specialist","Sign in as HR Specialist (hrspec). Open Performance → Rating scales and Review templates.","The lists are visible but 'New scale…' / 'New template…' / Edit are unavailable (write is HR Admin only)."),
 ("SEC-A2","Permissions (Phase A)","Employee","Sign in as a plain Employee. Try opening /performance/kpis and /performance/okrs directly.","The employee cannot see other employees' KPI assignments or OKR admin data (read is HR + managers; the API returns no rows / access denied for out-of-scope employees)."),
 ("SEC-A3","Permissions (Phase A)","Manager","Sign in as a Manager (manager). Open Performance → KPIs → Assignments.","The manager can view and assign/measure KPIs (managers are allowed to manage their team's KPIs); employee lists respect the manager's hierarchy scope."),
 # ---------- Phase B.1: goal-plan approval + progress trail (M392) ----------
 ("GPL-01","Goal plan submit (M392)","HR Admin","Performance → Goals. Pick a cycle + ONE employee (who has a manager). Create 2 DRAFT goals with weights 60 and 30. Look at the 'weights' tag next to the filters.","The tag shows 'weights 90%' in red and 'Submit plan for approval' is disabled (weights must total exactly 100 — §37.5)."),
 ("GPL-02","Goal plan submit (M392)","HR Admin","Edit one goal's weight so the two total 100 (60+40). Click 'Submit plan for approval'.","The tag turns green at 100%; submission succeeds — both goals show Approval = PENDING APPROVAL and editing them is now blocked."),
 ("GPL-03","Goal plan approve (M392)","Manager","Sign in as the employee's MANAGER → Approvals inbox. Find 'Goal plan: <employee> (2 goals, weights 100%)' and Approve it.","After approval the goals' Approval column shows APPROVED and their status flips DRAFT → ACTIVE."),
 ("GPL-04","Goal plan reject (M392)","HR Admin / Manager","Create another employee's plan (weights = 100), submit, and as their manager REJECT it in the Approvals inbox.","Goals show Approval = REJECTED and stay DRAFT — they can be edited and resubmitted (rework loop)."),
 ("GPL-05","Approved-edit guard (M392)","HR Admin","Edit an APPROVED goal (change its title or weight). Save, then look at its Approval column.","The edit works but Approval drops back to '—' (NOT_SUBMITTED) — a structurally changed goal must be resubmitted for approval (traceable)."),
 ("GPT-01","Progress trail (M392)","HR Admin","On any ACTIVE goal click 'Progress', set 40% with note 'Q2 checkpoint', save. Update again to 70% with another note. Then click 'History'.","The drawer shows BOTH updates newest-first: 40% → 70% with old→new values, notes, who and when (§6.4 audit trail)."),
 # ---------- Phase B.2: competency assessment (M393) ----------
 ("CMP-01","Competency seed (M393)","HR Admin","Open a performance review (Performance → Reviews → any review). In the 'Competency assessment' card click 'Seed from position'. (The employee's position must have required competencies — set them under Staffing → Position requirements if empty.)","One row per required competency appears with the Required level snapshot. If the position has none, a clear message says to add manually instead."),
 ("CMP-02","Competency add + assess (M393)","HR Admin","Click 'Add competency', pick one from the catalog with Required level 4. Then click 'Assess' on it: Self=3, Manager=3, Final=3, comment. Save.","The row shows levels 3/3/3 and Gap = '-1 below' in red (gap = required − final; positive gap = development need)."),
 ("CMP-03","Competency gap states (M393)","HR Admin","Assess another row with Final = required level, and one with Final ABOVE required.","Gap shows 'meets' (blue) when equal and '+N above' (green) when above."),
 ("CMP-04","Duplicate guard (M393)","HR Admin","Try adding the SAME competency to the review again.","Rejected — the competency is already on the review (it's also hidden from the picker)."),
 # ---------- Phase B.3: weighted scoring + override (M394) ----------
 ("SCR-01","Weighted scoring (M394)","HR Admin","On a review where goals are RATED (Goals page → Rate) and competencies have FINAL levels: in the 'Weighted scores (§18)' card enter Values score 4 and click 'Compute scores'.","Goal/KPI/Competency/Values cells populate; Overall = weighted average using the template weights (Goals 50 / KPI 25 / Competency 20 / Values 5), re-normalised over the sections that have scores; the §18.3 band label (e.g. 'Exceeds Expectations') appears next to it."),
 ("SCR-02","Missing-section handling (M394)","HR Admin","Compute scores on a review with NO KPI assignments (KPI cell —).","The overall still computes — the missing KPI section is left out and the remaining weights are re-normalised (it is NOT dragged down by a zero)."),
 ("OVR-01","HR override (M394)","HR Admin","On the scored review click 'Override rating'. New rating=4.5, Reason='Calibration committee decision'. Confirm.","Final rating becomes 4.50 with an 'overridden' tag; the Override cell shows the ORIGINAL computed rating, who overrode and the reason (original preserved — §18.4)."),
 ("OVR-02","Override guards (M394)","HR Admin / HR Specialist","Try overriding without a reason; then sign in as HR Specialist and look for the Override button.","No reason → rejected (reason is required). HR Specialist doesn't see the Override button (HR Admin only)."),
 ("SEC-B1","Permissions (Phase B)","Employee","Sign in as an Employee. Open another employee's review URL directly (/performance/reviews/<id>).","The competency and scoring data of another employee is NOT accessible (hierarchy guard — access denied / error)."),
 # ---------- Phase C: 360° nominations + questionnaires (M395) ----------
 ("Q36-01","Questionnaires (M395)","HR Admin","Performance → 360° nominations → Questionnaires tab → 'New questionnaire'. Code=STD_360, Name='Standard 360'. Add 3 questions: a RATING/COMPETENCY, a RATING/LEADERSHIP and a TEXT/IMPROVEMENT one. Save.","The questionnaire appears with 3 questions (expand the row to see types/categories/required tags)."),
 ("N36-01","Nomination (M395)","HR Admin","Nominations tab → pick a cycle → 'Nominate reviewer'. Subject=employee A, Reviewer=employee B, Relationship=PEER, Questionnaire=STD_360, Anonymous ON. Nominate.","A NOMINATED row appears with the PEER tag, questionnaire name and 'anonymous' tag."),
 ("N36-02","Duplicate guard (M395)","HR Admin","Nominate the SAME reviewer for the SAME subject in the SAME cycle again.","Rejected — this reviewer is already nominated for the subject in this cycle."),
 ("N36-03","Approve + respond (M395)","HR Admin","On the NOMINATED row click Approve, then Respond. The STD_360 questions render (stars for RATING, text area for TEXT). Answer required questions, set Overall rating 4, strengths/improvements. Submit.","Status becomes COMPLETED. Performance → Feedback shows a new ANONYMOUS PEER feedback for subject A (answers stored under competencies)."),
 ("N36-04","Decline (M395)","HR Admin","Nominate another reviewer, then Decline with a reason.","Status becomes DECLINED and the reason shows on the row."),
 ("N36-05","Min/max rules (M395)","HR Admin","Edit the review cycle: set Max reviewers=2 and Min reviewers=2 (cycle form). Then try nominating a 3rd live reviewer for the same subject.","Rejected — maximum reviewers reached. The summary tags next to the filters show 'completed / min' in red until 2 responses are completed."),
 # ---------- Phase D.1: acknowledgement + appeals (M396) ----------
 ("ACK-01","Acknowledgement (M396)","HR Admin","Open a review that HAS a final rating. In 'Acknowledgement & appeals' click 'Acknowledge result' with a comment, dispute OFF.","The card shows 'Acknowledged <time>' with a green 'accepted' tag; the Acknowledge button disappears (one-shot)."),
 ("ACK-02","Ack guards (M396)","HR Admin","Try acknowledging the same review again (or a review with NO final rating).","Already acknowledged → rejected; no final rating → the button is not offered / rejected."),
 ("APL-01","Appeal happy path (M396)","HR Admin","On a review with a final rating click 'Submit appeal' with a reason. Then (as HR) 'Take under review' → 'Decide' → APPROVED with Adjusted rating 4.2 and notes.","Appeal goes SUBMITTED → UNDER_REVIEW → APPROVED. The review's final rating becomes 4.20 with the 'overridden' tag and the ORIGINAL rating preserved in the Weighted-scores card (§37.12). Close the appeal → CLOSED."),
 ("APL-02","Appeal reject/return (M396)","HR Admin","Submit an appeal on another review; decide REJECTED (rating unchanged). Submit a third; decide RETURNED, then click Resubmit.","REJECTED keeps the rating; RETURNED lets the employee resubmit (status back to SUBMITTED). Only ONE live appeal per review is allowed."),
 # ---------- Phase D.2: calibration committee + outliers (M397) ----------
 ("COM-01","Committee (M397)","HR Admin","Performance → Review cycles → open a cycle's Calibration page. On a session click 'Committee'. Add a manager as MEMBER and someone as OBSERVER.","Both appear with role tags. (CHAIR/MEMBER may calibrate; OBSERVER is read-only.)"),
 ("COM-02","Committee gate (M397)","Manager","Sign in as a manager NOT on any committee and open the calibration board URL for the cycle.","Access denied — the board requires committee membership for managers (HR roles always may)."),
 ("COM-03","Committee calibrate (M397)","Manager","Add the manager to the committee as MEMBER (as HR), start the session, then as that manager calibrate a review on the board.","The calibrate saves — committee MEMBERs may calibrate while the session is IN_PROGRESS."),
 ("VIS-01","Notes visibility (M397)","Employee","As an employee whose review has calibration notes, open your own review.","Calibration notes (and any override reason) are NOT visible to you — they are HR/committee-only (§19.3). Signing in as HR shows them."),
 ("OUT-01","Outliers (M397)","HR Admin","On the Calibration page scroll to 'Outliers & manager leniency'.","Shows the cycle average, a per-manager average with Δ tags (orange/red when ≥0.5 off), and individual reviews ≥1.0 from the cycle average."),
 # ---------- Phase E.1: PIP (M398) ----------
 ("PIP-01","PIP create (M398)","HR Admin","Performance → PIPs → 'New PIP'. Pick an employee, reason, start today, end +60 days, objectives + success criteria. Save.","A DRAFT PIP appears. Trying to create a SECOND PIP for the same employee is rejected (one live PIP per employee)."),
 ("PIP-02","PIP lifecycle (M398)","HR Admin","Open the PIP → Activate. Then 'Acknowledge (employee)' with a comment. Then record a checkpoint (rating 2, comment).","DRAFT → ACTIVE (employee is notified) → ACKNOWLEDGED → after the checkpoint IN_PROGRESS. Checkpoints are blocked until acknowledged; the timeline shows the 2/5 rating in red."),
 ("PIP-03","PIP extend + close (M398)","HR Admin","Extend the PIP with a later end date; then 'Close with outcome' = IMPROVED with notes.","EXTENDED shows the new window; closing with IMPROVED → COMPLETED_SUCCESS (any other outcome → FAILED). Outcome decisions are HR-only; history stays (nothing is deleted)."),
 ("PIP-04","PIP evidence (M398)","HR Admin","In the PIP drawer upload a file under 'Evidence / documents'.","The file attaches via the standard uploader (M16 registry) and reloads with the PIP."),
 # ---------- Phase E.2: development plans (M399) ----------
 ("DVP-01","Dev plan (M399)","HR Admin","Performance → Development plans → 'New plan'. Employee, title, 3 actions of different types (e.g. Training course / Coaching / Stretch assignment). Save.","The plan appears at 0%; expanding shows the 3 typed actions."),
 ("DVP-02","Dev plan progress (M399)","HR Admin","Mark one action Done, one In progress.","Progress becomes 33.33% and status flips to IN PROGRESS. Marking ALL actions Done → 100% and the plan auto-COMPLETES."),
 ("DVP-03","From-gaps plan (M399)","HR Admin","On a review with positive competency gaps call the from-review generation (POST /api/performance/dev-plans/from-review/<reviewId> — or verify via the API tool).","A plan 'Close competency gaps — review …' is created with one TRAINING_COURSE action per positive gap (§21.3)."),
 # ---------- Phase E.3: continuous feedback + check-ins (M400) ----------
 ("CFB-01","Feedback (M400)","HR Admin","Performance → Check-ins → pick an employee → 'Give feedback': PRAISE, employee-visible, tags 'teamwork'. Then a second one: kind NOTE, visibility 'Private manager note'.","Both appear in the Feedback timeline; the private one carries a red 'manager-private' tag."),
 ("CFB-02","Private-note visibility (M400)","Employee","Sign in as that employee and load your feedback (via My Workspace/API).","The PRAISE note is visible; the MANAGER_PRIVATE note is NOT returned to the employee (§22.1)."),
 ("CIN-01","Check-in (M400)","HR Admin","'Record check-in': today, One-to-one, discussion notes + action items + follow-up date. Save. Then click Acknowledge on the row.","The check-in row appears with notes/actions/follow-up; acknowledging turns the Ack column into a green 'acknowledged' tag (one-shot)."),
 # ---------- Phase F: dashboard + notifications + comp guard (M401/M402) ----------
 ("DSH-01","Dashboard (M401)","HR Admin","Performance → Dashboard → pick the test cycle.","Stat cards populate (reviews, pending acknowledgement, disputed, goal plans awaiting approval, open appeals, active PIPs, active dev plans) plus the status funnel, rating distribution and top/bottom-5 performers."),
 ("NTF-01","Notifications (M402)","Employee","After a manager approves/rejects your goal plan, HR decides your appeal, or a PIP is activated for you — check the notifications bell.","In-app notifications arrive for each event (goal plan approved/rejected, appeal decision incl. adjusted rating, PIP activated)."),
 ("CMP-G1","Comp-bridge guard (M402)","HR Admin","Create a bonus run (Comp & Benefits → Bonus runs) for a cycle where one review is APPROVED/COMPLETED and another is still CALIBRATING.","Only the FINALISED (APPROVED/COMPLETED) reviews produce bonus items — the CALIBRATING review is excluded (pre-final ratings never reach compensation)."),
]

wb = Workbook()

# Instructions
ws = wb.active; ws.title = "Instructions"; ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 26; ws.column_dimensions["C"].width = 100
def put(r,l,v,bold=True):
    b=ws.cell(r,2,l); b.font=Font(name=FONT,bold=bold,size=11,color=navy if bold else "000000"); b.alignment=Alignment(vertical="top",wrap_text=True)
    c=ws.cell(r,3,v); c.font=Font(name=FONT,size=11); c.alignment=Alignment(vertical="top",wrap_text=True)
ws.cell(1,2,"Performance Management — User Acceptance Test (UAT) Script").font=Font(name=FONT,bold=True,size=16,color=navy)
ws.cell(2,2,"HCM_12 Performance Management  •  front-end (browser) testing  •  grows as each phase ships").font=Font(name=FONT,size=11,color=grey)
for r,l,v in [
 (4,"How to use","Open the 'Test Cases' sheet. Do exactly what the Steps say (which tab, which button, what to type), compare to the Expected Result, and pick Pass / Fail / Blocked / Not Run in the Result column. Add anything unusual under Tester Notes. Fill the Sign-off sheet at the end."),
 (6,"Application URL","http://localhost:5180 — sign in first. Performance features are under the top-nav 'Performance' menu."),
 (7,"Delivered so far","ALL PHASES A–F COMPLETE. A: rating scales (M388), review templates + cycle scoping (M389), KPI library + auto-scored assignments (M390), OKR + check-ins (M391). B: goal-plan approval workflow + progress trail (M392), competency assessment with gaps (M393), §18 weighted scoring + band + HR override (M394). C: 360° nominations + questionnaires + min/max rules (M395). D: acknowledgement + appeals (M396), calibration committee + HR-only notes + outliers (M397). E: PIPs (M398), development plans (M399), continuous feedback + check-ins (M400). F: HR dashboard (M401), notifications + comp-bridge finalised-only guard (M402)."),
 (9,"Logins you will need","HR Admin (full performance admin), HR Specialist (read-only checks), Manager (team KPI/OKR), Employee (self-service scoping checks). If you only have an admin login, mark role-restriction rows Blocked with a note."),
 (11,"Result values","Pass = worked as expected.  Fail = did not match (add a note).  Blocked = could not run (missing login/data).  Not Run = skipped."),
 (12,"Good to know","Rating-scale bands convert numeric scores to ratings during weighted scoring (Phase B). Templates fix which sections a review has and their weights (scoring sections must total 100). KPI ratings compute automatically: LINEAR = achievement ÷ 20 (100% → 5.0); THRESHOLD = bands ≥110→5 / ≥100→4 / ≥80→3 / ≥60→2 / else 1. OKR objective progress = weighted average of key-result progress."),
]:
    put(r,l,v)
ws.row_dimensions[4].height=45; ws.row_dimensions[7].height=60; ws.row_dimensions[12].height=60

# Test Cases
tc = wb.create_sheet("Test Cases"); tc.sheet_view.showGridLines=False
headers=["Test ID","Feature Area","Login As","Test Steps (what to click / type)","Expected Result","Result","Tester Notes"]
widths=[12,22,18,62,58,12,30]
for i,(h,w) in enumerate(zip(headers,widths),1):
    c=tc.cell(1,i,h); c.font=Font(name=FONT,bold=True,color=white,size=11); c.fill=PatternFill("solid",fgColor=navy)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
    tc.column_dimensions[get_column_letter(i)].width=w
tc.freeze_panes="A2"; tc.row_dimensions[1].height=26
for idx,(tid,feat,role,steps,exp) in enumerate(CASES):
    r=idx+2; banded=(idx%2==1)
    for col,v in enumerate([tid,feat,role,steps,exp,"",""],1):
        c=tc.cell(r,col,v)
        c.font=Font(name=FONT,size=11,bold=(col==1),color=(navy if col==1 else "000000"))
        c.alignment=Alignment(vertical="top",wrap_text=True,horizontal=("center" if col in(1,6) else "left"))
        c.border=border
        if banded and col!=6: c.fill=PatternFill("solid",fgColor=band)
    tc.cell(r,6).fill=PatternFill("solid",fgColor="FFF7E6")
    longest=max(len(steps),len(exp)); tc.row_dimensions[r].height=max(30,min(160,(longest//60+1)*15+12))
dv=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True)
dv.add(f"F2:F{len(CASES)+1}"); tc.add_data_validation(dv)

# Sign-off
so=wb.create_sheet("Sign-off"); so.sheet_view.showGridLines=False
for col,w in zip("ABCDE",[3,36,16,22,18]): so.column_dimensions[col].width=w
so.cell(1,2,"Performance UAT — Sign-off").font=Font(name=FONT,bold=True,size=15,color=navy)
for i,h in enumerate(["Feature Area","Result","Tester","Date"],2):
    c=so.cell(3,i,h); c.font=Font(name=FONT,bold=True,color=white); c.fill=PatternFill("solid",fgColor=blue); c.border=border
areas=["Rating scales (SCL)","Review templates + cycle scoping (TPL)",
       "KPI library + assignments + scoring (KPI)","OKR objectives + key results + check-ins (OKR)",
       "Goal-plan approval + progress trail (GPL/GPT)","Competency assessment (CMP)",
       "Weighted scoring + override (SCR/OVR)","360° nominations + questionnaires (Q36/N36)",
       "Acknowledgement + appeals (ACK/APL)","Calibration committee + outliers (COM/VIS/OUT)",
       "PIPs (PIP)","Development plans (DVP)","Continuous feedback + check-ins (CFB/CIN)",
       "Dashboard + notifications + comp guard (DSH/NTF/CMP-G)","Permissions (SEC)"]
dv2=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True); so.add_data_validation(dv2)
for j,a in enumerate(areas):
    r=4+j; so.cell(r,2,a).font=Font(name=FONT,size=11)
    for col in range(2,6): so.cell(r,col).border=border
    so.cell(r,3).fill=PatternFill("solid",fgColor="FFF7E6"); dv2.add(so.cell(r,3))
fr=4+len(areas)+2
so.cell(fr,2,"Overall decision (SHIP / DON'T SHIP):").font=Font(name=FONT,bold=True,size=12,color=navy)
so.cell(fr,3).fill=PatternFill("solid",fgColor="FFF7E6"); so.cell(fr,3).border=border
so.cell(fr+2,2,"Tester name:").font=Font(name=FONT,size=11); so.cell(fr+3,2,"Date:").font=Font(name=FONT,size=11)

wb.save(OUT)
print("WROTE",OUT,"cases:",len(CASES))
