"""Generator for the HCM_23-28 bundle (HRSD/DocMgmt/Workflow/ER/EHS/T&E) UAT Excel
test script. Append (id, feature, role, steps, expected) tuples to CASES as milestones
ship, then re-run to regenerate the workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/Users/elshad/Millers HCM/PRD/HCM_23_to_28_HRSD_DocMgmt_Workflow_ER_EHS_TravelExpense_Multi_Tenant_PRDs/HCM23_28_UAT_Test_Script.xlsx"
FONT = "Arial"
navy = "1F3864"; blue = "2E5496"; band = "F2F5FB"; grey = "808080"; white = "FFFFFF"
thin = Side(style="thin", color="BFBFBF"); border = Border(thin, thin, thin, thin)

CASES = [
 ("SET-01","Setup","HR Admin","Sign in as HR Admin at http://localhost:5180. Check the nav: Knowledge Base, the HR area (Service queue, Announcements, Document categories, Signatures), Approvals (Workflow definitions, Approval groups), the new Employee Relations group, the new Health & Safety group, and Business trips.","All the listed entries are present."),
 # ---------- 23 HRSD ----------
 ("KB-01","Knowledge base (M436)","HR Admin","Knowledge Base → create an article (title, category, markdown body), save as DRAFT, then Publish. Sign in as an employee and search for it by a keyword from the title.","Draft invisible to employees; after publish the employee finds it via search, views it (view count increments) and can vote helpful."),
 ("TCK-01","Ticket comments (M437)","Employee / HR","As an employee open one of your HR requests and add a comment. As HR reply with a normal comment AND an internal note.","The employee sees their comment + HR's reply but NOT the internal note; HR sees all three (internal visually distinct)."),
 ("QUE-01","Agent queues (M438)","HR Admin","Submit requests with categories PAYROLL_INQUIRY and POLICY_QUESTION, then open the HR service queue.","Each request auto-routed to its queue (PAYROLL / POLICY tabs show counts); reassigning to another queue works; category SLA applies (GRIEVANCE=1 business day, SALARY_CERT=2, POLICY_QUESTION=5)."),
 # ---------- 24 DocMgmt ----------
 ("DOC-01","Doc categories + versions (M439)","HR Admin","HR → Document categories: create/edit a category. Then on an employee's documents tab upload a document with that category, then use 'New version' to replace it.","Category saves and appears in the upload form; the new version becomes current and the history expand shows the version chain."),
 ("SIG-01","Signatures (M440)","HR Admin / Employee","HR → Signatures → create a request targeting an employee document with 1-2 signers. As a signer open My Workspace.","Signers see a 'Pending signatures' card; signing records timestamp; when all sign the request flips COMPLETED; a decline records the reason."),
 ("REN-01","Expiry renewal (M441)","HR Admin","Set a category to auto-renewal, give an employee document in it an expiry within 30 days, run/wait for the daily sweep (06:30) or trigger via the scheduler if exposed.","A DOCUMENT_RENEWAL HR service request is auto-created referencing the document; running again does NOT duplicate it."),
 # ---------- 25 Workflow ----------
 ("WFV-01","Definition versioning (M442)","HR Admin","Approvals → Workflow definitions.","Definitions list with code, version (existing rows = v1), effective window and step counts; read-only."),
 ("WFG-01","Approval groups (M443)","HR Admin","Approvals → Approval groups → create a group, add 2 usernames as members.","Group saves; a workflow step configured with this group lets ANY member act on the instance (verify via an inbox item if a def uses groups; otherwise verify CRUD + members only and mark partially tested)."),
 ("WFE-01","Auto-escalation (M444)","HR Admin","Configure a workflow step's escalation action to AUTO_APPROVE with a short SLA (or verify via an existing SLA-breached instance).","On SLA breach the instance is auto-approved by actor 'system-sla' with the audit comment 'Auto-actioned on SLA breach'."),
 # ---------- 26 ER (CONFIDENTIAL) ----------
 ("ERC-01","ER case (M445)","HR Admin","Employee Relations → ER Cases → create a COMPLAINT case (employee, severity HIGH). Add a note (internal), open an investigation, add an interview and an evidence file. Move status to UNDER_INVESTIGATION → RESOLVED → CLOSED.","Case gets an ER-number; every transition works and is audited; evidence uploads via the attachment control; closing works only after investigation handling."),
 ("ERC-02","Confidential case (M445)","HR Admin / HR Specialist","Create a case with Confidential ON, owner = HR Admin. Sign in as HR Specialist and open ER Cases.","The confidential case is INVISIBLE to the HR Specialist (list + direct URL); non-confidential cases remain visible. Legal-hold ON blocks closing."),
 ("ERC-03","Anonymous complaint (M445)","HR Admin","Create a case with NO employee selected (anonymous checkbox).","The case saves without an employee reference."),
 ("WRN-01","Warnings (M446)","HR Admin / Employee","Issue a FIRST_WRITTEN warning to an employee (note the auto-expiry hint +12 months). Sign in as that employee → My Workspace.","The employee sees a highlighted 'My warnings' entry and can Acknowledge (one-shot); HR list shows acknowledged tag; another employee sees nothing."),
 ("CAP-01","Corrective actions (M447)","HR Admin","Employee Relations → Corrective Actions → create a plan with a due date in the past (or wait for the 05:45 sweep). Transition another plan OPEN→IN_PROGRESS→COMPLETED.","The overdue plan shows/becomes OVERDUE (red); transitions work and are audited."),
 # ---------- 27 EHS ----------
 ("INC-01","Incident report (M448)","Employee","My Workspace → 'Report safety incident' (or Health & Safety → Incidents): report a NEAR_MISS (MODERATE) with one witness.","The incident gets an INC-number; the employee sees their own report in the list but NOT other employees' incidents."),
 ("INC-02","Critical incident (M448)","HR Admin","Report an INJURY incident with severity SERIOUS.","Investigation-required is auto-set and a notification fires; HR sees it in the full list and can close it only with a resolution."),
 ("INJ-01","Injury + RTW (M449)","HR Admin","On the SERIOUS incident add an injury report (body part, lost-time days). Then create a return-to-work plan; approve as the employee's manager, then as HR.","Injury data saves (HR-only visibility); the RTW plan shows both approval flags and activates; a manager outside the hierarchy cannot approve."),
 ("RSK-01","Risk register (M450)","HR Admin","Health & Safety → Risk register → create an assessment with likelihood 4 × impact 4.","Score 16 shows a red HIGH band (≥15); approving stamps approver; lowering to 2×2 shows LOW green (<5)."),
 ("INS-01","Inspections (M450)","HR Admin","Create an inspection with 4 findings; mark 3 OK and 1 NON_COMPLIANT, then Complete.","Overall score computes to 75%; the non-compliant finding can link to a corrective action."),
 ("ECA-01","EHS corrective actions (M451)","HR Admin","Create an EHS corrective action linked to the inspection with a past due date (or wait for the 05:50 sweep).","It appears in the queue and shows/turns OVERDUE red; completing sets closed timestamp."),
 ("PPE-01","PPE (M451)","HR Admin / Employee","Health & Safety → PPE assignments → issue a Helmet to an employee (expiry auto ≈ +24 months). Check the 'Expiring soon' tab with a short-expiry item (Mask +1 month). As the employee open My Workspace.","Issue auto-computes expiry from the catalog; the Mask appears under expiring-soon; the employee's 'My PPE' card lists their items; Return records condition."),
 # ---------- 28 T&E ----------
 ("PDM-01","Per-diem rules (M452)","HR Admin","Business trips → per-diem rules admin: check the seeds (Baku 30/25/5, Istanbul 40/35/5, Dubai 50/40/10...). Create a business trip to Istanbul for 3 days.","The trip form previews the per-diem breakdown (3 × 80 = 240 AZN with meals/lodging/incidentals split); most specific rule wins (city over country)."),
 ("MIL-01","Mileage claims (M453)","Employee / Manager","As an employee submit a mileage claim: personal car, 120 km.","Total auto-computes 120 × 0.30 = 36 AZN with an MC-number; the employee's MANAGER (or HR) approves — the employee cannot approve their own; HR can mark PAID."),
 ("EXP-01","Expense policy (M454)","HR Admin","Check the seeded expense policies. Submit an expense claim with a meal line over the daily limit (e.g. 80 AZN vs 50 limit) and a line above the 20-AZN receipt threshold without a receipt.","The over-limit line records a WARNING flag (claim still submits); the no-receipt line flags RECEIPT_REQUIRED; a BLOCKED category line rejects the claim outright."),
 ("RMB-01","Reimbursement batch (M455)","HR Admin","Create a reimbursement batch from APPROVED expense claims, approve it, then Mark paid with a payment reference.","Batch gets an RB-number and totals correctly; a claim can't join two live batches; on PAID each claim bridges a one-time EXPENSE_REIMBURSEMENT item toward payroll (visible in payroll bonuses/pending items) and flips to reimbursed."),
 # ---------- Cross-cutting ----------
 ("SEC-01","Confidentiality sweep","HR Specialist / Employee / Manager","(a) HR Specialist: try confidential ER cases and GRIEVANCE tickets. (b) Employee: try /er/cases, another employee's HR request, another's incident, injuries data. (c) Manager: try approving a mileage claim of someone outside your team.","All denied/empty: confidential ER + grievances are HR-Admin-only; employees reach only their own tickets/incidents/PPE; injuries are HR-only; approvals respect hierarchy."),
 ("SEC-02","No hard deletes","HR Admin","Look for delete buttons on ER cases, warnings, incidents, expense claims.","Nothing offers physical deletion — records close/archive only (legal-hold blocks even closing)."),
]

wb = Workbook()
ws = wb.active; ws.title = "Instructions"; ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 26; ws.column_dimensions["C"].width = 100
def put(r,l,v):
    b=ws.cell(r,2,l); b.font=Font(name=FONT,bold=True,size=11,color=navy); b.alignment=Alignment(vertical="top",wrap_text=True)
    c=ws.cell(r,3,v); c.font=Font(name=FONT,size=11); c.alignment=Alignment(vertical="top",wrap_text=True)
ws.cell(1,2,"HRSD / Documents / Workflow / ER / EHS / Travel-Expense (HCM_23-28) — UAT Script").font=Font(name=FONT,bold=True,size=16,color=navy)
ws.cell(2,2,"HCM_23-28 bundle  •  front-end (browser) testing").font=Font(name=FONT,size=11,color=grey)
for r,l,v in [
 (4,"How to use","Open 'Test Cases'. Follow the Steps exactly, compare to Expected Result, pick Pass / Fail / Blocked / Not Run. Note anything odd in Tester Notes. Fill the Sign-off sheet."),
 (6,"Application URL","http://localhost:5180 — sign in first."),
 (7,"Delivered","23 HRSD: knowledge base (M436), ticket comments (M437), agent queues + category SLA (M438). 24 Documents: categories + versioning (M439), signature requests (M440), expiry→renewal automation (M441). 25 Workflow: definition versioning (M442), approval groups (M443), SLA auto-approve/reject (M444). 26 ER: confidential case management + investigations (M445), warning ladder + acknowledgement (M446), corrective actions (M447). 27 EHS: incidents (M448), injuries + return-to-work (M449), risk register + inspections (M450), corrective actions + PPE (M451). 28 T&E: per-diem rules (M452), mileage claims (M453), expense policy engine (M454), reimbursement batches + payroll bridge (M455)."),
 (9,"Logins you will need","HR Admin, HR Specialist (confidentiality checks), Manager (hierarchy/approval checks), Employee (ESS + scoping). Mark role-restricted rows Blocked if a login is missing."),
 (11,"Result values","Pass / Fail (note!) / Blocked / Not Run."),
 (12,"CONFIDENTIALITY under test","(1) Confidential ER cases = HR-Admin + case-owner only; grievance tickets HR-Admin only. (2) Injury/medical data HR-Admin only. (3) Employees see only their own tickets/incidents/warnings/PPE. (4) No physical deletes anywhere; legal hold blocks closure. Any leak = automatic Fail + report immediately."),
]:
    put(r,l,v)
ws.row_dimensions[7].height=90; ws.row_dimensions[12].height=60

tc = wb.create_sheet("Test Cases"); tc.sheet_view.showGridLines=False
headers=["Test ID","Feature Area","Login As","Test Steps (what to click / type)","Expected Result","Result","Tester Notes"]
widths=[12,26,20,62,58,12,30]
for i,(h,w) in enumerate(zip(headers,widths),1):
    c=tc.cell(1,i,h); c.font=Font(name=FONT,bold=True,color=white,size=11); c.fill=PatternFill("solid",fgColor=navy)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
    tc.column_dimensions[get_column_letter(i)].width=w
tc.freeze_panes="A2"; tc.row_dimensions[1].height=26
for idx,(tid,feat,role,steps,exp) in enumerate(CASES):
    r=idx+2; banded=(idx%2==1)
    for col,v in enumerate([tid,feat,role,steps,exp,"",""],1):
        c=tc.cell(r,col,v)
        c.font=Font(name=FONT,size=11,bold=(col==1),color=(navy if col==1 else "000000"))
        c.alignment=Alignment(vertical="top",wrap_text=True,horizontal=("center" if col in(1,6) else "left"))
        c.border=border
        if banded and col!=6: c.fill=PatternFill("solid",fgColor=band)
    tc.cell(r,6).fill=PatternFill("solid",fgColor="FFF7E6")
    longest=max(len(steps),len(exp)); tc.row_dimensions[r].height=max(30,min(160,(longest//60+1)*15+12))
dv=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True)
dv.add(f"F2:F{len(CASES)+1}"); tc.add_data_validation(dv)

so=wb.create_sheet("Sign-off"); so.sheet_view.showGridLines=False
for col,w in zip("ABCDE",[3,44,16,22,18]): so.column_dimensions[col].width=w
so.cell(1,2,"HCM_23-28 bundle UAT — Sign-off").font=Font(name=FONT,bold=True,size=15,color=navy)
for i,h in enumerate(["Feature Area","Result","Tester","Date"],2):
    c=so.cell(3,i,h); c.font=Font(name=FONT,bold=True,color=white); c.fill=PatternFill("solid",fgColor=blue); c.border=border
areas=["Knowledge base (KB)","Ticket comments (TCK)","Agent queues + SLA (QUE)",
       "Document categories + versions (DOC)","Signatures (SIG)","Expiry renewal (REN)",
       "Workflow versioning/groups/escalation (WFV/WFG/WFE)",
       "ER cases + confidentiality (ERC)","Warnings (WRN)","ER corrective actions (CAP)",
       "Incidents (INC)","Injury + return-to-work (INJ)","Risk register + inspections (RSK/INS)",
       "EHS corrective actions + PPE (ECA/PPE)",
       "Per-diem (PDM)","Mileage (MIL)","Expense policy (EXP)","Reimbursement (RMB)",
       "Confidentiality + no-delete (SEC)"]
dv2=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True); so.add_data_validation(dv2)
for j,a in enumerate(areas):
    r=4+j; so.cell(r,2,a).font=Font(name=FONT,size=11)
    for col in range(2,6): so.cell(r,col).border=border
    so.cell(r,3).fill=PatternFill("solid",fgColor="FFF7E6"); dv2.add(so.cell(r,3))
fr=4+len(areas)+2
so.cell(fr,2,"Overall decision (SHIP / DON'T SHIP):").font=Font(name=FONT,bold=True,size=12,color=navy)
so.cell(fr,3).fill=PatternFill("solid",fgColor="FFF7E6"); so.cell(fr,3).border=border
so.cell(fr+2,2,"Tester name:").font=Font(name=FONT,size=11); so.cell(fr+3,2,"Date:").font=Font(name=FONT,size=11)

wb.save(OUT)
print("WROTE",OUT,"cases:",len(CASES))
