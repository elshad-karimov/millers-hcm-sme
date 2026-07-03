"""Generator for the Benefits Administration UAT Excel test script.
Extended per phase — append new (id, feature, role, steps, expected) tuples to CASES
as milestones are delivered, then re-run this script to regenerate the workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/Users/elshad/Millers HCM/PRD/HCM_11_Benefits_Administration_Multi_Tenant_PRD/Benefits_UAT_Test_Script.xlsx"
FONT = "Arial"
navy = "1F3864"; blue = "2E5496"; band = "F2F5FB"; grey = "808080"; white = "FFFFFF"
thin = Side(style="thin", color="BFBFBF"); border = Border(thin, thin, thin, thin)

# Delivered so far: Phase A — M373 categories, M374 providers, M375 plan tiers + eligibility.
# Phases B–F append here as they ship.
CASES = [
 ("SET-01","Setup","HR Admin","Sign in as HR Admin at http://localhost:5180. Open the top-nav 'Benefits' page (Benefits administration). You should see tabs: Categories, Providers, Plans catalog, Enrolments, My benefits.","The Benefits page opens with the five tabs visible."),
 # ---------- Phase A.1: categories (M373) ----------
 ("CAT-01","Categories (M373)","HR Admin","Benefits → Categories tab.","Eight seeded categories are listed in order: Health Insurance, Life Insurance, Pension / Retirement, Meal Allowance, Transport Allowance, Housing Allowance, Mobile / Communication, Wellness Program."),
 ("CAT-02","Categories (M373)","HR Admin","On the Categories list, read the Taxable and Provider columns for Health vs Meal.","Health = Exempt + Provider Required; Meal = Taxable + no provider. (Insured plans are tax-exempt; cash-like allowances are taxable.)"),
 ("CAT-03","Categories (M373)","HR Admin","Click 'New category…'. Code=GYM, Name=Gym Membership, tick Taxable off, Provider off, Order=9. Save.","The new category GYM appears at the end of the list, Active."),
 ("CAT-04","Categories (M373)","HR Admin","Click 'New category…' again and re-use the code HEALTH. Save.","Rejected — category code must be unique; an error is shown and no row is created."),
 ("CAT-05","Categories (M373)","HR Admin","Open the GYM category, untick Active, Save. Then toggle the 'Active only' switch.","GYM shows Inactive (not deleted); with 'Active only' on it is hidden from the list."),
 # ---------- Phase A.2: providers (M374) ----------
 ("PRV-01","Providers (M374)","HR Admin","Benefits → Providers tab → 'New provider…'. Code=PASHA-INS, Name=Pasha Insurance OJSC, Type=Insurer, Contact email=account@pasha.az, Contract number=C-2026-01, Contract window = today → +1 year. Save.","The provider PASHA-INS appears in the table with type Insurer, the contract number and the contract window."),
 ("PRV-02","Providers (M374)","HR Admin","Create a provider with a Contract window whose end date is BEFORE its start date. Save.","Rejected — contract end must be on/after start; an error is shown."),
 ("PRV-03","Providers (M374)","HR Admin","Create a second provider (Code=CAPITAL-PF, Name=Kapital Pension Fund, Type=Pension fund). Save.","The pension-fund provider appears."),
 ("PRV-04","Providers (M374)","HR Admin","Enter an invalid email (e.g. 'abc') in Contact email and try to save.","Rejected — the email format is validated."),
 # ---------- Phase A.3: plans + tiers + eligibility (M375) ----------
 ("PLN-01","Plans (M375)","HR Admin","Benefits → Plans catalog → 'New plan…'. Code=HEALTH-FAM-26, Name=Family Health 2026, Type=Health, Category=HEALTH — Health Insurance, Plan year=2026, Provider (from master)=Pasha Insurance, Employer/mo=120, Employee/mo=30, Effective window=this year. Save.","The plan appears in the table showing the Category tag (Health Insurance), Year 2026 and Provider = Pasha Insurance OJSC (the master name, not free text)."),
 ("PLN-02","Plans (M375)","HR Admin","On the HEALTH-FAM-26 row click 'Tiers…'. In the Coverage tiers section click 'Add tier' three times and set: Employee only Er=100/Ee=0; Employee + spouse Er=150/Ee=40; Family Er=220/Ee=90 (add a Cover/sum-insured on one). Save.","The three tiers save; reopening 'Tiers…' shows them with the right employer/employee amounts."),
 ("PLN-03","Plans (M375)","HR Admin","Reopen 'Tiers…' and add a second tier with the SAME tier code as an existing one (e.g. two 'Family'). Save.","Rejected — each coverage tier code must be unique (message shown)."),
 ("PLN-04","Plans (M375)","HR Admin","In 'Tiers…' → Eligibility rules, click 'Add rule'. Employment type=FULL_TIME, Min service=3 (mo), Note='FT after probation'. Save.","The eligibility rule saves; reopening shows it. (No rules = open to all; this rule limits the plan to full-time staff with 3+ months service — enforced at enrolment in Phase B.)"),
 ("PLN-05","Plans (M375)","HR Admin","Open the plan editor for HEALTH-FAM-26 and confirm Category and Plan year are populated; change Provider (from master) to blank and instead type a free-text Provider. Save, then look at the Provider column.","With a master provider it shows the master name; cleared, it falls back to the free-text provider name."),
 # ---------- Phase A permissions ----------
 ("SEC-A1","Permissions (Phase A)","HR Specialist","Sign in as HR Specialist. Open Benefits → Categories / Providers / Plans.","You can view the lists, but 'New…' / Save / the Tiers… editor are read-only or unavailable (write is HR Admin / Benefits Manager only)."),
 ("SEC-A2","Permissions (Phase A)","Employee","Sign in as a plain Employee and open the Benefits page.","Only the 'My benefits' tab is available — the admin tabs (Categories, Providers, Plans, Enrolments) are hidden."),
 # ---------- Phase B.1: enrolment overhaul — tiers + dependants (M376) ----------
 ("SETB","Setup B","HR Admin","Pick a test employee who has at least one benefit-eligible dependant (Employee profile → Dependants → make sure 'Benefit eligible' or 'Insurance eligible' is ticked on one). Copy the employee's UUID from their profile URL. Make sure the HEALTH-FAM-26 plan from Phase A has coverage tiers (Tiers… editor).","The employee has an eligible dependant and you have their UUID; the plan has tiers."),
 ("ENR-01","Enrolments (M376)","HR Admin","Benefits → Enrolments → 'Enrol employee…'. Pick the plan; paste the employee UUID (eligible dependants load below); pick a Coverage tier (e.g. Employee + spouse); tick a dependant; Start date today; Status = Enrolled. Enrol.","A new ENROLLED row appears showing the Tier tag, the covered-dependant count (hover shows the name) and the Employee/mo amount taken from the chosen tier (not the plan flat rate)."),
 ("ENR-02","Enrolments (M376)","HR Admin","Enrol the same or another employee with NO tier selected.","The enrolment uses the plan's flat employer/employee contribution; Tier column shows 'flat'."),
 ("ENR-03","Enrolments (M376)","HR Admin","Try to enrol and tick a dependant who is NOT benefit/insurance eligible (or belongs to a different employee).","Rejected — only eligible dependants of that employee can be covered (error shown). Ineligible dependants don't even appear in the list."),
 # ---------- Phase B.2: enrolment approval workflow (M377) ----------
 ("APR-01","Approvals (M377)","HR Admin","Enrol an employee but set Status = 'Draft (submit for approval)'. Then on that DRAFT row click Submit.","Status goes DRAFT → PENDING_APPROVAL and the request appears in the Approvals inbox."),
 ("APR-02","Approvals (M377)","HR Admin","Open the Approvals inbox and approve the pending benefit enrolment.","After approval the enrolment status becomes ENROLLED (coverage now active)."),
 ("APR-03","Approvals (M377)","HR Admin","Submit another draft enrolment, then reject it in the Approvals inbox.","The enrolment status becomes REJECTED (coverage never activates)."),
 ("APR-04","Approvals (M377)","HR Admin","On a DRAFT (not yet submitted) or PENDING enrolment click Cancel.","Status becomes CANCELLED; nothing activates."),
 ("SUS-01","Suspend/Resume (M376)","HR Admin","On an ENROLLED row click Suspend and confirm; then click Resume.","Status toggles ENROLLED → SUSPENDED → ENROLLED. (Suspending pauses the payroll deduction; resuming restarts it — see DED-03.)"),
 # ---------- Phase B.3: payroll deduction bridge (M378) — PAYROLL-IMPACTING ----------
 ("DED-01","Payroll bridge (M378)","HR Admin","Enrol an employee (Status Enrolled) in a plan whose Employee/mo contribution is > 0 (e.g. 30). Then open Payroll → Deductions for that employee (or the employee's payslip preview).","A RECURRING payroll deduction 'Benefit contribution: <plan>' for the employee amount (e.g. 30.00) is present and ACTIVE, starting the enrolment month."),
 ("DED-02","Payroll bridge (M378)","HR Admin","Enrol an employee in an employer-paid plan (Employee/mo = 0).","NO payroll deduction is created (nothing to deduct)."),
 ("DED-03","Payroll bridge (M378)","HR Admin","Terminate (or Suspend) the DED-01 enrolment, then re-check that employee's deductions.","The benefit-contribution deduction is now CANCELLED (not deleted). Re-running payroll for a later month no longer deducts it."),
 ("DED-04","Payroll bridge (M378)","Payroll Specialist","Run payroll for a period where the DED-01 employee is enrolled; open their payslip.","Net pay is reduced by the benefit employee-contribution (pre-tax: it lowers taxable gross before tax/DSMF/MMI). The deduction shows in the payslip breakdown."),
 ("SEC-B1","Permissions (Phase B)","HR Specialist","Sign in as HR Specialist and open Benefits → Enrolments.","You can view enrolments but cannot enrol / submit / approve / suspend / terminate (write is HR Admin / Benefits Manager only)."),
]

wb = Workbook()

# Instructions
ws = wb.active; ws.title = "Instructions"; ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 26; ws.column_dimensions["C"].width = 100
def put(r,l,v,bold=True):
    b=ws.cell(r,2,l); b.font=Font(name=FONT,bold=bold,size=11,color=navy if bold else "000000"); b.alignment=Alignment(vertical="top",wrap_text=True)
    c=ws.cell(r,3,v); c.font=Font(name=FONT,size=11); c.alignment=Alignment(vertical="top",wrap_text=True)
ws.cell(1,2,"Benefits Administration — User Acceptance Test (UAT) Script").font=Font(name=FONT,bold=True,size=16,color=navy)
ws.cell(2,2,"HCM_11 Benefits Administration  •  front-end (browser) testing  •  grows as each phase ships").font=Font(name=FONT,size=11,color=grey)
for r,l,v in [
 (4,"How to use","Open the 'Test Cases' sheet. Do exactly what the Steps say (which tab, which button, what to type), compare to the Expected Result, and pick Pass / Fail / Blocked / Not Run in the Result column. Add anything unusual under Tester Notes. Fill the Sign-off sheet at the end."),
 (6,"Application URL","http://localhost:5180 — sign in first. Benefits features are under the top-nav 'Benefits' page."),
 (7,"Delivered so far","Phases A–B COMPLETE. A: Categories (M373), Providers (M374), Plans + coverage tiers + eligibility rules (M375). B: enrolment with tiers + dependant-level coverage (M376), enrolment approval workflow (M377), and the payroll deduction bridge — employee contributions flow to payroll as recurring pre-tax deductions (M378). Open enrolment, life events, claims, dashboards and letters arrive in later phases."),
 (9,"Logins you will need","HR Admin (full benefits access), HR Specialist (read-only), Payroll Specialist (payroll runs), Employee (self-service). If you only have an admin login, mark the role-restriction rows Blocked with a note."),
 (11,"Result values","Pass = worked as expected.  Fail = did not match (add a note).  Blocked = could not run (missing login/data).  Not Run = skipped."),
 (12,"Good to know","Categories drive tax treatment + whether a plan needs a provider. Plans link to a category, a plan year and a provider from the master. Coverage tiers set the employer/employee split per coverage level; eligibility rules decide who may enrol (enforced from Phase B)."),
]:
    put(r,l,v)
ws.row_dimensions[4].height=45; ws.row_dimensions[7].height=60; ws.row_dimensions[12].height=45

# Test Cases
tc = wb.create_sheet("Test Cases"); tc.sheet_view.showGridLines=False
headers=["Test ID","Feature Area","Login As","Test Steps (what to click / type)","Expected Result","Result","Tester Notes"]
widths=[12,22,18,62,58,12,30]
for i,(h,w) in enumerate(zip(headers,widths),1):
    c=tc.cell(1,i,h); c.font=Font(name=FONT,bold=True,color=white,size=11); c.fill=PatternFill("solid",fgColor=navy)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
    tc.column_dimensions[get_column_letter(i)].width=w
tc.freeze_panes="A2"; tc.row_dimensions[1].height=26
for idx,(tid,feat,role,steps,exp) in enumerate(CASES):
    r=idx+2; banded=(idx%2==1)
    for col,v in enumerate([tid,feat,role,steps,exp,"",""],1):
        c=tc.cell(r,col,v)
        c.font=Font(name=FONT,size=11,bold=(col==1),color=(navy if col==1 else "000000"))
        c.alignment=Alignment(vertical="top",wrap_text=True,horizontal=("center" if col in(1,6) else "left"))
        c.border=border
        if banded and col!=6: c.fill=PatternFill("solid",fgColor=band)
    tc.cell(r,6).fill=PatternFill("solid",fgColor="FFF7E6")
    longest=max(len(steps),len(exp)); tc.row_dimensions[r].height=max(30,min(160,(longest//60+1)*15+12))
dv=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True)
dv.add(f"F2:F{len(CASES)+1}"); tc.add_data_validation(dv)

# Sign-off
so=wb.create_sheet("Sign-off"); so.sheet_view.showGridLines=False
for col,w in zip("ABCDE",[3,36,16,22,18]): so.column_dimensions[col].width=w
so.cell(1,2,"Benefits UAT — Sign-off").font=Font(name=FONT,bold=True,size=15,color=navy)
for i,h in enumerate(["Feature Area","Result","Tester","Date"],2):
    c=so.cell(3,i,h); c.font=Font(name=FONT,bold=True,color=white); c.fill=PatternFill("solid",fgColor=blue); c.border=border
areas=["Categories (CAT)","Providers (PRV)","Plans + Tiers + Eligibility (PLN)",
       "Enrolment: tiers + dependants (ENR)","Enrolment approval (APR)","Suspend / Resume (SUS)",
       "Payroll deduction bridge (DED)","Permissions (SEC)"]
dv2=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True); so.add_data_validation(dv2)
for j,a in enumerate(areas):
    r=4+j; so.cell(r,2,a).font=Font(name=FONT,size=11)
    for col in range(2,6): so.cell(r,col).border=border
    so.cell(r,3).fill=PatternFill("solid",fgColor="FFF7E6"); dv2.add(so.cell(r,3))
fr=4+len(areas)+2
so.cell(fr,2,"Overall decision (SHIP / DON'T SHIP):").font=Font(name=FONT,bold=True,size=12,color=navy)
so.cell(fr,3).fill=PatternFill("solid",fgColor="FFF7E6"); so.cell(fr,3).border=border
so.cell(fr+2,2,"Tester name:").font=Font(name=FONT,size=11); so.cell(fr+3,2,"Date:").font=Font(name=FONT,size=11)

wb.save(OUT)
print("WROTE",OUT,"cases:",len(CASES))
