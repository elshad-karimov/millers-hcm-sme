"""Generator for the HCM_34-49 bundle UAT Excel test script (M481-M494 gaps over the
16-module PRD; pre-built modules are covered by prior bundle UAT packs)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/Users/elshad/Millers HCM/PRD/HCM_34_to_49_Engagement_Rewards_Scheduling_Costing_GlobalMobility_Security_Integration_Multi_Tenant_PRDs/HCM34_49_UAT_Test_Script.xlsx"
FONT = "Arial"
navy = "1F3864"; blue = "2E5496"; band = "F2F5FB"; grey = "808080"; white = "FFFFFF"
thin = Side(style="thin", color="BFBFBF"); border = Border(thin, thin, thin, thin)

CASES = [
 ("SET-01","Setup","HR Admin","Sign in at http://localhost:5180. Check nav: Engagement (Reward Points), roster tabs (Open Shifts, Swap Requests), Timesheets (Projects), Payroll (Labor Rates), Reports (Labor Cost), new Mobility and Contingent Workforce groups, Policies (Campaigns), Admin (Integrations, Notification Templates, Permission Matrix).","All entries present."),
 # Rewards (M481/M483)
 ("RWD-01","Reward catalog + grant (M481)","HR Admin","Engagement → Reward Points: create a catalog item (100 points, monetary value 100 AZN, taxable). Create a budget for this year. Grant 200 points to an employee with a note.","Item + budget save; grant succeeds and decrements the budget's remaining amount; over-budget grants are rejected."),
 ("RWD-02","Redeem + fulfil (M481)","Employee / HR Admin","As the employee (My Workspace → My Reward Points): balance shows 200; redeem the 100-point item. As HR fulfil the redemption (optionally against a payroll run).","Wallet drops to 100; redemption goes REQUESTED→FULFILLED; a ONE_TIME bonus row appears for monetary items; insufficient-balance redemption is rejected."),
 ("RWD-03","Anniversary recognition (M483)","HR Admin","Set/verify an employee whose hire anniversary is today (or trigger the daily job) and open the recognition wall.","A service-anniversary recognition appears once (no duplicate on re-run); toggle off via the tenant setting stops new ones."),
 # Scheduling (M482)
 ("SCH-01","Open shifts (M482)","HR Admin / Employee","Roster → Open Shifts: publish an open shift with 1 slot. As an employee claim it. As a second employee try claiming.","Claim creates the roster entry and fills the slot; the second claim is rejected (FILLED)."),
 ("SCH-02","Shift swap (M482)","Employee / Manager","As employee A request a swap of one of your roster entries to employee B. As A's MANAGER approve it. Try approving a swap you requested yourself.","On approval the roster entry moves to B (both audited); self-approval is blocked; a manager outside the hierarchy cannot approve."),
 # Projects + costing (M484-M486)
 ("PTS-01","Projects + day tagging (M484)","HR Admin","Timesheets → Projects: create project PRJ-1 (billing rate). Open an employee's timesheet and tag a day with PRJ-1 + task code + billable.","Project saves; the day stores project/task/billable (if day-editing UI is absent, mark Blocked and note)."),
 ("LBC-01","Labor rates (M485)","HR Admin","Payroll → Labor Rates: create an hourly rate with an effective window.","Rate saves; only Finance/HR roles can open the page."),
 ("LBC-02","Cost allocation + report (M485/M486)","HR Admin","Approve a timesheet containing tagged days, then open Reports → Labor Cost for that month.","Allocations compute (hours × rate) and the by-project / by-department tabs show matching totals; a MANAGER (non-Finance) cannot open the report."),
 # Mobility (M487)
 ("MOB-01","Assignment tracking (M487)","HR Admin","Mobility → International Assignments: create an assignment (host country/city, dates, visa expiry within 90 days, allowance fields). Upload a document in the drawer. Check the Expiring Visas tab.","Assignment gets an IA-number; document attaches; the visa appears in expiring list with a red tag; status transitions audited."),
 ("MOB-02","Assignment privacy (M487)","Employee","As the assigned employee view your own assignment (if surfaced); as ANOTHER employee try the assignment by id/API.","Own assignment readable; others get denied — assignments are HR + own-employee only."),
 # Contingent (M488/M489)
 ("CNT-01","Contractor engagement (M488)","HR Admin","Contingent → Contractors: create an engagement for a CONTRACTOR-type employee (rate, unit, PO, contract end soon).","Engagement saves; tenure alert appears near contract end; rate/PO visible to HR/Finance only (manager surfaces show none)."),
 ("CNT-02","Convert to FTE (M489)","HR Admin","On the ACTIVE engagement use Convert to FTE (employment type + effective date).","Employee's employment type flips; engagement becomes CONVERTED with the date; audited; contractor never appeared in statutory payroll before conversion."),
 # Policy campaigns (M490)
 ("POL-01","Re-ack campaign (M490)","HR Admin / Employee","Policies → Campaigns: create a campaign for a published policy version (audience ALL, due +14d), Launch it. As an employee acknowledge the policy. Check campaign progress.","Launch notifies the audience; progress shows acked/total climbing and per-department rows; Close ends it."),
 # Movement execution (M491)
 ("MOV-01","Execute movement (M491)","HR Admin","On an APPROVED transfer/promotion request click Execute and confirm.","Employee's org unit/position actually changes; request shows EXECUTED with executedAt/By; the confirm modal notes salary changes go through the salary-change flow; executing a non-APPROVED request is blocked."),
 # Integration registry (M492)
 ("INT-01","Integration registry (M492)","System Admin","Admin → Integrations: create a config (type WEBHOOK, endpoint URL, credentials reference name). Check the logs drawer and failures tab.","Config saves (endpoint stored encrypted, never echoed raw in lists if masked); logs/failures render; page denied to non-SYSTEM_ADMIN (HR Admin cannot open)."),
 # Notification templates (M493)
 ("NTF-01","Templates + delivery logs (M493)","HR Admin","Admin → Notification Templates: create an EMAIL template with {{name}} variables. Open the Delivery Logs tab after some notifications fire.","Template saves and renders variables; delivery logs list SENT/FAILED rows with filters (if no rows yet, verify empty state loads — note it)."),
 # Permission matrix (M494)
 ("SEC-01","Permission matrix (M494)","System Admin / Employee","Admin → Permission Matrix. Then as a plain employee try the page/API.","Roles × capability grid renders with colored levels; employees are denied."),
 # Cross-cutting
 ("SEC-02","Money confidentiality sweep","Manager / Employee","(a) Manager: look for contractor rates/PO, labor rates/costs, reward budgets on any accessible surface. (b) Employee: try another employee's wallet / redemptions / assignment / engagement by id.","All absent/denied: money data is HR/Finance-only; wallets and assignments are self-or-HR."),
 ("SEC-03","Payroll boundary","HR Admin","Verify reward fulfilment and cost allocation created only bonus/allocation rows (no payroll run executed or changed by any bundle action).","No payroll run status changed; bonuses appear as pending one-time items for the next run only."),
]

wb = Workbook()
ws = wb.active; ws.title = "Instructions"; ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 26; ws.column_dimensions["C"].width = 100
def put(r,l,v):
    b=ws.cell(r,2,l); b.font=Font(name=FONT,bold=True,size=11,color=navy); b.alignment=Alignment(vertical="top",wrap_text=True)
    c=ws.cell(r,3,v); c.font=Font(name=FONT,size=11); c.alignment=Alignment(vertical="top",wrap_text=True)
ws.cell(1,2,"HCM_34-49 (16 modules) — UAT Script for the new gap milestones M481-M494").font=Font(name=FONT,bold=True,size=15,color=navy)
ws.cell(2,2,"Front-end (browser) testing • pre-existing module functionality is covered by earlier bundle UAT packs").font=Font(name=FONT,size=11,color=grey)
for r,l,v in [
 (4,"How to use","Open 'Test Cases'. Follow Steps exactly, compare to Expected Result, pick Pass / Fail / Blocked / Not Run. Note oddities in Tester Notes. Fill the Sign-off sheet."),
 (6,"Application URL","http://localhost:5180 — sign in first."),
 (7,"Delivered (new work)","Rewards: catalog + points wallet + budgets + redemption→bonus bridge (M481), anniversary auto-recognition (M483). Scheduling: open shifts + swap requests (M482). Project timesheets (M484) + labor rates/cost allocation (M485) + cost reports (M486). Global mobility assignment tracking (M487). Contractor engagements (M488) + FTE conversion (M489). Policy re-ack campaigns (M490). Movement execution (M491). Integration registry (M492). Notification templates + delivery logs (M493). Permission matrix (M494). Modules 34/41/42/43/45/46/47 were already ≥85% built — covered by earlier UAT packs."),
 (9,"Logins","HR Admin, System Admin (integrations/matrix), Manager, Employee. Mark role-restricted rows Blocked if a login is missing."),
 (11,"Result values","Pass / Fail (note!) / Blocked / Not Run."),
 (12,"KEY RULES under test","(1) Money data (contractor rates/PO, labor costs, reward budgets) = HR/Finance-only. (2) Wallets/assignments/redemptions = self-or-HR. (3) Nothing executes payroll — only instruction/bonus/allocation rows. (4) Self-approval blocked on swaps. Any breach = automatic Fail + report."),
]:
    put(r,l,v)
ws.row_dimensions[7].height=90; ws.row_dimensions[12].height=60

tc = wb.create_sheet("Test Cases"); tc.sheet_view.showGridLines=False
headers=["Test ID","Feature Area","Login As","Test Steps (what to click / type)","Expected Result","Result","Tester Notes"]
widths=[12,28,22,62,58,12,30]
for i,(h,w) in enumerate(zip(headers,widths),1):
    c=tc.cell(1,i,h); c.font=Font(name=FONT,bold=True,color=white,size=11); c.fill=PatternFill("solid",fgColor=navy)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
    tc.column_dimensions[get_column_letter(i)].width=w
tc.freeze_panes="A2"; tc.row_dimensions[1].height=26
for idx,(tid,feat,role,steps,exp) in enumerate(CASES):
    r=idx+2; banded=(idx%2==1)
    for col,v in enumerate([tid,feat,role,steps,exp,"",""],1):
        c=tc.cell(r,col,v)
        c.font=Font(name=FONT,size=11,bold=(col==1),color=(navy if col==1 else "000000"))
        c.alignment=Alignment(vertical="top",wrap_text=True,horizontal=("center" if col in(1,6) else "left"))
        c.border=border
        if banded and col!=6: c.fill=PatternFill("solid",fgColor=band)
    tc.cell(r,6).fill=PatternFill("solid",fgColor="FFF7E6")
    longest=max(len(steps),len(exp)); tc.row_dimensions[r].height=max(30,min(160,(longest//60+1)*15+12))
dv=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True)
dv.add(f"F2:F{len(CASES)+1}"); tc.add_data_validation(dv)

so=wb.create_sheet("Sign-off"); so.sheet_view.showGridLines=False
for col,w in zip("ABCDE",[3,44,16,22,18]): so.column_dimensions[col].width=w
so.cell(1,2,"HCM_34-49 bundle UAT — Sign-off").font=Font(name=FONT,bold=True,size=15,color=navy)
for i,h in enumerate(["Feature Area","Result","Tester","Date"],2):
    c=so.cell(3,i,h); c.font=Font(name=FONT,bold=True,color=white); c.fill=PatternFill("solid",fgColor=blue); c.border=border
areas=["Reward catalog/wallet/budgets (RWD-01/02)","Anniversary recognition (RWD-03)",
       "Open shifts + swaps (SCH)","Projects + day tagging (PTS)","Labor rates + costing (LBC)",
       "International assignments (MOB)","Contractors + conversion (CNT)",
       "Policy campaigns (POL)","Movement execution (MOV)","Integration registry (INT)",
       "Notification templates + logs (NTF)","Permission matrix (SEC-01)",
       "Confidentiality + payroll boundary (SEC-02/03)"]
dv2=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True); so.add_data_validation(dv2)
for j,a in enumerate(areas):
    r=4+j; so.cell(r,2,a).font=Font(name=FONT,size=11)
    for col in range(2,6): so.cell(r,col).border=border
    so.cell(r,3).fill=PatternFill("solid",fgColor="FFF7E6"); dv2.add(so.cell(r,3))
fr=4+len(areas)+2
so.cell(fr,2,"Overall decision (SHIP / DON'T SHIP):").font=Font(name=FONT,bold=True,size=12,color=navy)
so.cell(fr,3).fill=PatternFill("solid",fgColor="FFF7E6"); so.cell(fr,3).border=border
so.cell(fr+2,2,"Tester name:").font=Font(name=FONT,size=11); so.cell(fr+3,2,"Date:").font=Font(name=FONT,size=11)

wb.save(OUT)
print("WROTE",OUT,"cases:",len(CASES))
