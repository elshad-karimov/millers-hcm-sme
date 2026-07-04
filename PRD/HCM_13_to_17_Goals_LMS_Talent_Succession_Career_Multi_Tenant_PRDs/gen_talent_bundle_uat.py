"""Generator for the HCM_13-17 bundle (Goals/LMS/Talent/Succession/Career) UAT Excel
test script. Append (id, feature, role, steps, expected) tuples to CASES as milestones
ship, then re-run to regenerate the workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/Users/elshad/Millers HCM/PRD/HCM_13_to_17_Goals_LMS_Talent_Succession_Career_Multi_Tenant_PRDs/Talent_Bundle_UAT_Test_Script.xlsx"
FONT = "Arial"
navy = "1F3864"; blue = "2E5496"; band = "F2F5FB"; grey = "808080"; white = "FFFFFF"
thin = Side(style="thin", color="BFBFBF"); border = Border(thin, thin, thin, thin)

CASES = [
 ("SET-01","Setup","HR Admin","Sign in as HR Admin at http://localhost:5180. Check the top nav: Performance menu (Goals), Learning menu (Classroom training, Mandatory training) and the new Talent menu (Pools, Reviews, Analytics, Succession plans, Replacement chart, Career paths, Mentoring, Career dashboard).","All the listed menu entries are present."),
 # ---------- HCM_13: goal types (M403) ----------
 ("GT-01","Goal types (M403)","HR Admin","Performance → Goals → open the 'New goal' form and expand the Goal type dropdown.","14 seeded goal types are offered (Revenue, Sales, Cost optimisation, Project delivery, Quality, Customer satisfaction, Operational excellence, Compliance, Safety, Innovation, Leadership, Team development, Personal development, Process improvement)."),
 ("GT-02","Goal types (M403)","HR Admin","Create a goal picking Type='Personal development'. Watch the Category field as you pick the type. Save and find the goal in the list.","Picking the type auto-selects its category (DEVELOPMENT); the saved goal's row shows the type as a tag in the Type column."),
 # ---------- HCM_14: instructors + rooms (M404) ----------
 ("INS-01","Instructors (M404)","HR Admin","Learning → Classroom training → Instructors tab → 'New instructor'. Pick an internal employee, specialization 'Leadership', hourly cost 50. Save. Then create a SECOND instructor with NO employee but External name 'ACME Training Ltd'.","Both instructors appear — one shows the employee's name (internal), the other the external name. An instructor with neither employee nor external name is rejected."),
 ("ROM-01","Rooms (M404)","HR Admin","Rooms tab → 'New room'. Code=TR-1, Name='Training room 1', Capacity=2. Save. Create a second room marked Virtual.","Both rooms appear; the virtual one carries a Virtual tag."),
 # ---------- HCM_14: sessions + attendance (M405) ----------
 ("SES-01","Sessions (M405)","HR Admin","Sessions tab → 'New session'. Pick a published course, instructor, room TR-1 (capacity 2), today 10:00→12:00. Save.","The session appears as SCHEDULED with course/instructor/room columns filled. An end time before the start time is rejected."),
 ("SES-02","Capacity (M405)","HR Admin","Open the session's attendance drawer and enrol 2 employees. Then try enrolling a 3rd.","First two enrol OK (ENROLLED). The 3rd is rejected — room TR-1 capacity (2) is full. Enrolling the SAME employee twice is also rejected."),
 ("SES-03","Attendance marking (M405)","HR Admin","Mark one enrolled employee ATTENDED and leave the other ENROLLED. Then set the session status to COMPLETED.","The ATTENDED row gets a check-in timestamp; on COMPLETED the still-ENROLLED employee is auto-finalised to NO_SHOW."),
 ("SES-04","Edit guard (M405)","HR Admin","Try editing the now-COMPLETED session (change its time).","Rejected — only SCHEDULED sessions can be edited."),
 # ---------- HCM_14: mandatory training (M406) ----------
 ("MND-01","Mandatory rule (M406)","HR Admin","Learning → Mandatory training → 'New rule'. Pick a published course, audience = one department, recurrence 12 months, due days 30. Save.","The rule appears, Active, with the audience and recurrence shown."),
 ("MND-02","Sweep (M406)","HR Admin","Click 'Run sweep now'. Then check Learning → My learning (or the course's enrolments) for an employee in that department.","The sweep reports how many were enrolled; in-scope ACTIVE employees now have an ASSIGNED enrolment with a due date ~30 days out. Running the sweep again enrols 0 (idempotent)."),
 ("MND-03","Compliance (M406)","HR Admin","Open the rule's compliance view.","Per-rule counts show in-scope / compliant / pending / overdue that match the enrolments just created."),
 # ---------- HCM_14: costs + feedback (M407/M408) ----------
 ("CST-01","Training costs (M407)","HR Admin","On the classroom page open the costs surface for the completed session (or course). Add cost lines: INSTRUCTOR 100 AZN and CATERING 40 AZN.","Both lines save and list with a total; a negative amount is rejected."),
 ("FBK-01","Feedback (M408)","HR Admin","Submit training feedback for the session as the ATTENDED employee (ratings 1–5: content 5, instructor 4, overall 5; anonymous ON). Then try submitting feedback for the NO_SHOW employee.","The attendee's feedback saves (one per employee per session). The non-attendee is rejected — feedback requires attendance."),
 ("FBK-02","Anonymous + instructor rating (M408)","HR Admin","View the session's feedback as HR and check the instructor's row on the Instructors tab.","The anonymous feedback shows NO employee identity; the instructor's average rating updated from the submitted instructor score."),
 # ---------- HCM_15: talent profile + interests (M409) ----------
 ("TPR-01","Talent profile (M409)","HR Admin","Talent → (profiles surface) → upsert a profile for employee A: HiPo ON (Manual), Retention risk HIGH, risk reason, retention action, mobility INTERNAL, aspirations.","The profile saves and lists employee A with HiPo + HIGH risk tags. Filtering by HiPo-only / risk works."),
 ("TPR-02","Career interests (M409)","Employee","Sign in as an employee. Via My Workspace / career dashboard, add a career interest: target role 'Team Lead', timeline 1–2 years. Then delete it and re-add.","The interest saves, lists and deletes — employees manage their OWN interests. A blank target role is rejected."),
 ("SEC-15A","CONFIDENTIALITY (M409)","Employee","As employee A (who has the HiPo/HIGH-risk profile), call GET /api/talent/profiles and /api/talent/profiles/employees/<own id> (or try the Talent pages).","Access denied — employees NEVER see talent profiles, including their own HiPo flag and retention risk. Talent pages/menu are hidden for plain employees."),
 # ---------- HCM_15: talent pools (M410) ----------
 ("POOL-01","Talent pools (M410)","HR Admin","Talent → Pools → 'New pool'. Code=LEAD_PIPE, Name='Leadership pipeline', purpose, review every 6 months. Save. Add 2 employees as members. Remove 1.","The pool saves; members add and list with who/when added; duplicate member add is rejected; removal works."),
 # ---------- HCM_15: talent review cycles (M411) ----------
 ("TRV-01","Talent review (M411)","HR Admin","Talent → Reviews → create cycle 'Talent review 2026'. Record a review for employee B: performance box 3, potential box 3, HiPo decision YES, retention risk MEDIUM, notes. Save.","The review saves (9/9 box). Recording AGAIN for the same employee in the same cycle UPDATES the existing review (upsert, no duplicate row)."),
 ("TRV-02","Profile roll-up (M411)","HR Admin","After TRV-01, open employee B's talent profile.","The profile now shows HiPo=YES with source NINE_BOX and retention risk MEDIUM — the review decision rolled onto the profile automatically."),
 # ---------- HCM_15: talent analytics (M412) ----------
 ("TAN-01","Talent analytics (M412)","HR Admin","Talent → Analytics → pick the 2026 cycle.","The 9-box distribution reflects the recorded reviews; HiPo count/% matches; the retention-risk breakdown and pool coverage (pools + member counts) populate."),
 # ---------- HCM_16: critical positions + succession plans (M413) ----------
 ("SUC-01","Critical positions (M413)","HR Admin","Talent → Succession plans → Critical positions tab → mark a position critical: reason, replacement difficulty HIGH, vacancy risk CRITICAL, succession required.","The critical position is listed with its difficulty/risk tags."),
 ("SUC-02","Succession plan + approval (M413)","HR Admin","Plans tab → 'New plan' for that critical position: plan owner, effective/review dates, emergency successor. Save (DRAFT), then Submit for approval. Approve it from the Approvals inbox (HR).","DRAFT → SUBMITTED starts a SUCCESSION_PLAN_APPROVAL workflow; approving flips the plan to APPROVED; rejecting another plan returns it to DRAFT. Both transitions are audit-logged."),
 # ---------- HCM_16: risk of loss (M414) ----------
 ("RSK-01","Risk of loss (M414)","HR Admin","Open the succession nominations page (9-box/succession area) and edit a nomination: Risk of loss HIGH, Impact of loss CRITICAL, risk reason, retention action.","The nomination saves and its row shows the risk/impact tags; invalid values outside LOW/MEDIUM/HIGH/CRITICAL are rejected."),
 # ---------- HCM_16: replacement chart + coverage (M415) ----------
 ("CHT-01","Replacement chart (M415)","HR Admin","Talent → Replacement chart.","Each critical position shows its incumbent(s) and nominated successors with readiness + risk; coverage cards count positions with 0 / 1 / 2+ READY_NOW successors."),
 ("CHT-02","Dev-plan bridge (M415)","HR Admin","On a successor nomination, attach a development plan (create one under Performance → Development plans first if needed).","The nomination links to the dev plan; the link is visible on the chart/nomination row."),
 # ---------- HCM_17: career paths (M416) ----------
 ("CPT-01","Career paths (M416)","HR Admin","Talent → Career paths → 'New path'. Code=ENG_TRACK, Name='Engineering track', job family. Add 2 steps: step 1 from Junior → Senior (required skills, 2 years, courses); step 2 Senior → Lead. Save. Re-open and edit a step, save again.","The path saves with ordered steps; re-saving (full replace) does not error and keeps the edited steps. All signed-in users can VIEW paths; only HR can edit."),
 # ---------- HCM_17: mentoring (M417) ----------
 ("MNT-01","Mentor + request (M417)","HR Admin","Talent → Mentoring → register a mentor (employee M, areas 'Leadership', max mentees 1). Then as (or on behalf of) employee E request mentoring with mentor M.","The mentor registers; the request appears as REQUESTED with a default 6-month term."),
 ("MNT-02","Capacity + lifecycle (M417)","HR Admin","Approve the request (→ ACTIVE). Then have a SECOND mentee request mentor M and try to approve it too. Finally COMPLETE the first relationship.","The second approval is rejected — mentor M is at max capacity (1 active). Completing the first frees capacity; statuses flow REQUESTED→ACTIVE→COMPLETED (or CANCELLED)."),
 ("SEC-17A","Mentoring visibility (M417)","Employee","As an employee who is neither the mentor, the mentee nor HR, try to read the mentoring relationship / its meeting notes.","Denied — mentoring relationships and notes are visible only to the mentor, the mentee and HR."),
 # ---------- HCM_17: career dashboard + recommendations (M418) ----------
 ("CDB-01","Career dashboard (M418)","Employee","Sign in as an employee → Talent → Career dashboard (own record).","The dashboard shows the employee's own career interests, development plans and mentoring status. NO HiPo/retention/succession data appears anywhere on it."),
 ("CDB-02","Job recommendations (M418)","Employee","With a career interest whose target role matches an OPEN internal job posting's title, reload the dashboard.","The matching posting appears under recommendations with a match score (role-title match scores higher than department-only match)."),
 ("CDB-03","Scope guard (M418)","Manager","As a manager, open the career dashboard for one of YOUR reports, then try an employee OUTSIDE your hierarchy (by employeeId in the URL/API).","Your report's dashboard loads; the out-of-hierarchy employee is denied (access scope guard)."),
 # ---------- Cross-cutting permissions ----------
 ("SEC-B1","Permissions","HR Specialist","Sign in as HR Specialist. Open Talent pools / reviews / analytics / succession / replacement chart.","Readable (HR Specialist may VIEW talent + succession data) but create/edit actions are unavailable (writes = HR Admin only)."),
 ("SEC-B2","Permissions","Manager","Sign in as a Manager (not HR). Try the Talent pages (pools, reviews, analytics, succession plans).","Denied / hidden — talent-review, HiPo and succession data are HR-and-executive-only; managers see them only via HR-run sessions, not directly."),
 ("SEC-B3","Permissions","Employee","Sign in as a plain Employee. Try /talent/pools, /talent/reviews, /talent/succession-plans and the classroom/mandatory ADMIN pages directly by URL.","All denied or empty — no talent/succession/LMS-admin surface leaks to employees; only their own career dashboard and interests work."),
]

wb = Workbook()

# Instructions
ws = wb.active; ws.title = "Instructions"; ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 26; ws.column_dimensions["C"].width = 100
def put(r,l,v,bold=True):
    b=ws.cell(r,2,l); b.font=Font(name=FONT,bold=bold,size=11,color=navy if bold else "000000"); b.alignment=Alignment(vertical="top",wrap_text=True)
    c=ws.cell(r,3,v); c.font=Font(name=FONT,size=11); c.alignment=Alignment(vertical="top",wrap_text=True)
ws.cell(1,2,"Goals / LMS / Talent / Succession / Career (HCM_13-17) — UAT Script").font=Font(name=FONT,bold=True,size=16,color=navy)
ws.cell(2,2,"HCM_13-17 bundle  •  front-end (browser) testing  •  grows as each phase ships").font=Font(name=FONT,size=11,color=grey)
for r,l,v in [
 (4,"How to use","Open the 'Test Cases' sheet. Do exactly what the Steps say (which tab, which button, what to type), compare to the Expected Result, and pick Pass / Fail / Blocked / Not Run in the Result column. Add anything unusual under Tester Notes. Fill the Sign-off sheet at the end."),
 (6,"Application URL","http://localhost:5180 — sign in first. Features live under the Performance, Learning and Talent top-nav menus."),
 (7,"Delivered so far","ALL PHASES COMPLETE. HCM_13: goal-type catalog wired into goals (M403). HCM_14: instructors + rooms (M404), classroom sessions + capacity + attendance (M405), mandatory-training rules + auto-enrol sweep + compliance (M406), training costs (M407), post-training feedback + instructor rating (M408). HCM_15: talent profile HiPo/retention (M409), talent pools (M410), talent review cycles/9-box → profile roll-up (M411), talent analytics (M412). HCM_16: critical positions + succession plans + approval workflow (M413), risk/impact-of-loss (M414), replacement chart + coverage + dev-plan bridge (M415). HCM_17: career paths (M416), mentoring (M417), career dashboard + internal job recommendations (M418)."),
 (9,"Logins you will need","HR Admin (full admin), HR Specialist (read-only talent checks), Manager (hierarchy checks), Employee (self-service + confidentiality checks). If you only have an admin login, mark role-restriction rows Blocked with a note."),
 (11,"Result values","Pass = worked as expected.  Fail = did not match (add a note).  Blocked = could not run (missing login/data).  Not Run = skipped."),
 (12,"CONFIDENTIALITY rule under test","Employees must NEVER see their own (or anyone's) HiPo flag, retention risk, talent review or succession nomination — these are HR/executive-only. Several SEC cases verify exactly this; treat any leak as an automatic Fail + report immediately."),
]:
    put(r,l,v)
ws.row_dimensions[4].height=45; ws.row_dimensions[7].height=75; ws.row_dimensions[12].height=45

# Test Cases
tc = wb.create_sheet("Test Cases"); tc.sheet_view.showGridLines=False
headers=["Test ID","Feature Area","Login As","Test Steps (what to click / type)","Expected Result","Result","Tester Notes"]
widths=[12,24,18,62,58,12,30]
for i,(h,w) in enumerate(zip(headers,widths),1):
    c=tc.cell(1,i,h); c.font=Font(name=FONT,bold=True,color=white,size=11); c.fill=PatternFill("solid",fgColor=navy)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
    tc.column_dimensions[get_column_letter(i)].width=w
tc.freeze_panes="A2"; tc.row_dimensions[1].height=26
for idx,(tid,feat,role,steps,exp) in enumerate(CASES):
    r=idx+2; banded=(idx%2==1)
    for col,v in enumerate([tid,feat,role,steps,exp,"",""],1):
        c=tc.cell(r,col,v)
        c.font=Font(name=FONT,size=11,bold=(col==1),color=(navy if col==1 else "000000"))
        c.alignment=Alignment(vertical="top",wrap_text=True,horizontal=("center" if col in(1,6) else "left"))
        c.border=border
        if banded and col!=6: c.fill=PatternFill("solid",fgColor=band)
    tc.cell(r,6).fill=PatternFill("solid",fgColor="FFF7E6")
    longest=max(len(steps),len(exp)); tc.row_dimensions[r].height=max(30,min(160,(longest//60+1)*15+12))
dv=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True)
dv.add(f"F2:F{len(CASES)+1}"); tc.add_data_validation(dv)

# Sign-off
so=wb.create_sheet("Sign-off"); so.sheet_view.showGridLines=False
for col,w in zip("ABCDE",[3,40,16,22,18]): so.column_dimensions[col].width=w
so.cell(1,2,"HCM_13-17 bundle UAT — Sign-off").font=Font(name=FONT,bold=True,size=15,color=navy)
for i,h in enumerate(["Feature Area","Result","Tester","Date"],2):
    c=so.cell(3,i,h); c.font=Font(name=FONT,bold=True,color=white); c.fill=PatternFill("solid",fgColor=blue); c.border=border
areas=["Goal types (GT)","Instructors + rooms (INS/ROM)","Classroom sessions + attendance (SES)",
       "Mandatory training (MND)","Training costs + feedback (CST/FBK)",
       "Talent profile + interests (TPR)","Talent pools (POOL)","Talent reviews + roll-up (TRV)",
       "Talent analytics (TAN)","Critical positions + succession plans (SUC)",
       "Risk of loss (RSK)","Replacement chart + coverage (CHT)","Career paths (CPT)",
       "Mentoring (MNT)","Career dashboard + recommendations (CDB)","Permissions + confidentiality (SEC)"]
dv2=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True); so.add_data_validation(dv2)
for j,a in enumerate(areas):
    r=4+j; so.cell(r,2,a).font=Font(name=FONT,size=11)
    for col in range(2,6): so.cell(r,col).border=border
    so.cell(r,3).fill=PatternFill("solid",fgColor="FFF7E6"); dv2.add(so.cell(r,3))
fr=4+len(areas)+2
so.cell(fr,2,"Overall decision (SHIP / DON'T SHIP):").font=Font(name=FONT,bold=True,size=12,color=navy)
so.cell(fr,3).fill=PatternFill("solid",fgColor="FFF7E6"); so.cell(fr,3).border=border
so.cell(fr+2,2,"Tester name:").font=Font(name=FONT,size=11); so.cell(fr+3,2,"Date:").font=Font(name=FONT,size=11)

wb.save(OUT)
print("WROTE",OUT,"cases:",len(CASES))
