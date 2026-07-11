"""Generator for the HCM_50 Mobile HR UAT Excel test script — DEVICE testing of the
Flutter app (M495-M512 new work over the existing 11 screens)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/Users/elshad/Millers HCM/PRD/HCM_50_Mobile_HR_Multi_Tenant_PRD/HCM50_Mobile_UAT_Test_Script.xlsx"
FONT = "Arial"
navy = "1F3864"; blue = "2E5496"; band = "F2F5FB"; grey = "808080"; white = "FFFFFF"
thin = Side(style="thin", color="BFBFBF"); border = Border(thin, thin, thin, thin)

CASES = [
 ("SET-01","Setup","Any","Install the app on a device/emulator (flutter run or the built APK). Confirm the backend is reachable and Keycloak login works. Note the home screen tiles.","App launches, login via Keycloak succeeds, home shows tiles incl. new: Attendance, Announcements, Policies, HR Requests, Documents, Directory, plus a notification bell."),
 ("AUTH-01","Auth / biometric","Employee","Log out, log back in. Then background+foreground the app and confirm biometric unlock (if enabled in Settings).","Keycloak OIDC login works; biometric (Face/fingerprint) unlocks a stored session; PIN/device-credential fallback works."),
 ("ATT-01","Clock in/out (M496)","Employee","Attendance tile → grant location permission (consent notice shown) → Clock In. Then Clock Out.","Punch records; today-status card updates; punch history shows the entries; geofenceStatus (INSIDE/OUTSIDE/UNKNOWN) + flagged shown; location-recorded notice appeared."),
 ("ATT-02","Geofence (M496/M497)","Employee","With a work location configured with GPS coords + radius, clock in from inside vs (simulate) outside the radius.","Inside → INSIDE, not flagged. Outside → OUTSIDE + flagged (still recorded, routed for review). If no geofence configured, punch allowed without geofence."),
 ("ATT-03","Offline punch queue (M508)","Employee","Turn on airplane mode, Clock In (queues). See pending-sync count. Restore network, tap Sync now (or reopen the screen).","Offline punch queues with device timestamp; on reconnect it syncs to the server; no duplicate is created on re-sync (idempotent)."),
 ("ATT-04","Corrections (M498)","Employee","Attendance → Corrections → submit a missing-punch correction with a reason.","Correction submits to the backend and appears in the corrections list."),
 ("ANN-01","Announcements (M499)","Employee","Announcements tile → open an announcement.","Active announcements list (date/audience); tapping shows full body."),
 ("POL-01","Policy ack (M500)","Employee","Policies tile → open a published policy → Acknowledge.","Policy body shows; Acknowledge records a timestamp; the policy then shows as acknowledged."),
 ("HRR-01","HR requests (M501)","Employee","HR Requests tile → submit a ticket (category + subject + description). View it in My Requests; add a comment.","Ticket submits with an SR-number/status; comments post and display."),
 ("DOC-01","Document upload (M502)","Employee","Documents tile → pick a photo/file, choose a type, upload.","Upload shows progress and succeeds; the document appears in the list with its type."),
 ("PRO-01","Profile edit (M503)","Employee","Profile → Edit Profile → change phone / emergency contact → submit.","Each changed field submits as a change request (pending approval); bank/salary are not editable; pending state shown."),
 ("DIR-01","Directory (M504)","Employee","Directory tile → search a colleague → open their card → tap phone / email.","Search returns public fields only (name/dept/position/work email+phone); NO salary/bank/national-id/DOB/home address; tel:/mailto: launch."),
 ("NTF-01","Notifications center (M505)","Employee","Tap the bell → view in-app notifications → mark read → tap one that deep-links.","Notifications list with unread badge; mark-read works; tapping navigates to the related screen where supported."),
 ("SET-02","Settings (M506)","Employee","Settings → toggle language en↔az, toggle biometric, view app version/session.","Language toggle re-localizes date/time pickers; biometric toggle persists; version/session shown."),
 ("PAY-01","Payslip security (M509)","Employee","Payslips tab → confirm a biometric/PIN re-auth is required before amounts show. Try to screenshot (Android).","Amounts hidden until re-auth (5-min validity); Android blocks the screenshot (FLAG_SECURE); iOS is best-effort (documented)."),
 ("OFF-01","Offline read cache (M507)","Employee","Load profile / leave balances / announcements / policies online, then go offline and reopen them.","Cached data shows with an 'offline / cached' banner; payslip amounts are NOT available offline (never cached)."),
 ("TEAM-01","Team calendar (M510)","Manager","As a manager, open the Team Calendar tile.","Team leave/BT/holidays render; entries show 'On Leave' — NOT the specific leave type. A non-manager doesn't see the tile / gets an empty state."),
 ("LV-01","Leave enhancements (M511)","Employee","Leave → submit with half-day toggle + a replacement + an attachment. Then cancel a PENDING request.","Half-day/replacement/attachment all submit; a pending request can be cancelled by the employee (own only) and the balance restores."),
 ("APR-01","Bulk approvals (M512)","Manager","Approvals → multi-select mode → select 2+ items → Approve selected (confirm).","Batch approve runs per item; per-item result summary shown; items you can't act on fail individually without blocking the rest."),
 ("SEC-01","Confidentiality / IDOR","Employee","Try (via the app or by crafting API calls) to read another employee's payslip, wallet, directory-private fields, or documents.","All denied/absent: the app only shows what the API returns for YOU; payslip/self endpoints scope to the logged-in employee; directory exposes public fields only."),
 ("SEC-02","No plaintext secrets","Tester/Dev","Inspect device storage / the offline cache after use.","Auth tokens are in secure storage (Keychain/Keystore), not plaintext; no salary/payslip/bank data in the shared-preferences cache; no secrets hardcoded in the app."),
]

wb = Workbook()
ws = wb.active; ws.title = "Instructions"; ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 26; ws.column_dimensions["C"].width = 100
def put(r,l,v):
    b=ws.cell(r,2,l); b.font=Font(name=FONT,bold=True,size=11,color=navy); b.alignment=Alignment(vertical="top",wrap_text=True)
    c=ws.cell(r,3,v); c.font=Font(name=FONT,size=11); c.alignment=Alignment(vertical="top",wrap_text=True)
ws.cell(1,2,"Mobile HR (HCM_50) — Device UAT Test Script").font=Font(name=FONT,bold=True,size=16,color=navy)
ws.cell(2,2,"Flutter app  •  on-device testing  •  new work M495-M512 over the existing screens").font=Font(name=FONT,size=11,color=grey)
for r,l,v in [
 (4,"How to use","Run the app on a phone or emulator. Do exactly what the Steps say, compare to the Expected Result, pick Pass / Fail / Blocked / Not Run. Note device model + OS in Tester Notes. Fill the Sign-off sheet."),
 (6,"App / backend","Flutter app in mobile/ (flutter run, or build an APK). Backend on the configured API base; Keycloak for login. Some features need device sensors (GPS, biometric, camera)."),
 (7,"Delivered (new)","Attendance clock-in + GPS/geofence (M496/M497) + offline punch queue (M508) + corrections (M498); announcements (M499); policy acknowledgement (M500); HR service requests (M501); document upload (M502); profile edit (M503); employee directory (M504); notifications center (M505); settings + language (M506); offline read cache (M507); payslip re-auth + screenshot block (M509); team calendar (M510); leave enhancements + self-cancel (M511); bulk approvals (M512). Existing screens (login/home/leave/payslip/timesheet/training/goals/team) already shipped earlier."),
 (9,"Logins","Employee (self-service + confidentiality), Manager (team calendar, bulk approvals). Mark manager-only rows Blocked if you only have an employee login."),
 (11,"Result values","Pass / Fail (add a note) / Blocked / Not Run."),
 (12,"KEY RULES under test","Thin client: the app shows only what the backend returns for YOU. Payslip amounts require re-auth and are never cached in plaintext. Directory = public fields only. Team calendar hides leave type. Location capture shows a consent notice. Tokens live in secure storage. Any leak of salary/bank/another person's data = automatic Fail + report."),
]:
    put(r,l,v)
ws.row_dimensions[7].height=90; ws.row_dimensions[12].height=60

tc = wb.create_sheet("Test Cases"); tc.sheet_view.showGridLines=False
headers=["Test ID","Feature Area","Login As","Test Steps (what to tap / do)","Expected Result","Result","Tester Notes"]
widths=[12,26,16,62,58,12,30]
for i,(h,w) in enumerate(zip(headers,widths),1):
    c=tc.cell(1,i,h); c.font=Font(name=FONT,bold=True,color=white,size=11); c.fill=PatternFill("solid",fgColor=navy)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
    tc.column_dimensions[get_column_letter(i)].width=w
tc.freeze_panes="A2"; tc.row_dimensions[1].height=26
for idx,(tid,feat,role,steps,exp) in enumerate(CASES):
    r=idx+2; banded=(idx%2==1)
    for col,v in enumerate([tid,feat,role,steps,exp,"",""],1):
        c=tc.cell(r,col,v)
        c.font=Font(name=FONT,size=11,bold=(col==1),color=(navy if col==1 else "000000"))
        c.alignment=Alignment(vertical="top",wrap_text=True,horizontal=("center" if col in(1,6) else "left"))
        c.border=border
        if banded and col!=6: c.fill=PatternFill("solid",fgColor=band)
    tc.cell(r,6).fill=PatternFill("solid",fgColor="FFF7E6")
    longest=max(len(steps),len(exp)); tc.row_dimensions[r].height=max(30,min(160,(longest//60+1)*15+12))
dv=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True)
dv.add(f"F2:F{len(CASES)+1}"); tc.add_data_validation(dv)

so=wb.create_sheet("Sign-off"); so.sheet_view.showGridLines=False
for col,w in zip("ABCDE",[3,44,16,22,18]): so.column_dimensions[col].width=w
so.cell(1,2,"Mobile HR UAT — Sign-off").font=Font(name=FONT,bold=True,size=15,color=navy)
for i,h in enumerate(["Feature Area","Result","Tester","Date"],2):
    c=so.cell(3,i,h); c.font=Font(name=FONT,bold=True,color=white); c.fill=PatternFill("solid",fgColor=blue); c.border=border
areas=["Auth / biometric (AUTH)","Attendance + geofence + offline (ATT)","Announcements (ANN)","Policy ack (POL)",
       "HR requests (HRR)","Document upload (DOC)","Profile edit (PRO)","Directory (DIR)",
       "Notifications (NTF)","Settings/language (SET-02)","Payslip security (PAY)","Offline cache (OFF)",
       "Team calendar (TEAM)","Leave enhancements (LV)","Bulk approvals (APR)","Confidentiality + secrets (SEC)"]
dv2=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True); so.add_data_validation(dv2)
for j,a in enumerate(areas):
    r=4+j; so.cell(r,2,a).font=Font(name=FONT,size=11)
    for col in range(2,6): so.cell(r,col).border=border
    so.cell(r,3).fill=PatternFill("solid",fgColor="FFF7E6"); dv2.add(so.cell(r,3))
fr=4+len(areas)+2
so.cell(fr,2,"Overall decision (SHIP / DON'T SHIP):").font=Font(name=FONT,bold=True,size=12,color=navy)
so.cell(fr,3).fill=PatternFill("solid",fgColor="FFF7E6"); so.cell(fr,3).border=border
so.cell(fr+2,2,"Tester name:").font=Font(name=FONT,size=11); so.cell(fr+3,2,"Device + OS:").font=Font(name=FONT,size=11); so.cell(fr+4,2,"Date:").font=Font(name=FONT,size=11)

wb.save(OUT)
print("WROTE",OUT,"cases:",len(CASES))
