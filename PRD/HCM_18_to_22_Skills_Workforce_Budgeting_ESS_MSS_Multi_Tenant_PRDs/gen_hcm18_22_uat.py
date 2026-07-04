"""Generator for the HCM_18-22 bundle (Skills/Workforce/Budgeting/ESS/MSS) UAT Excel
test script. Append (id, feature, role, steps, expected) tuples to CASES as milestones
ship, then re-run to regenerate the workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/Users/elshad/Millers HCM/PRD/HCM_18_to_22_Skills_Workforce_Budgeting_ESS_MSS_Multi_Tenant_PRDs/HCM18_22_UAT_Test_Script.xlsx"
FONT = "Arial"
navy = "1F3864"; blue = "2E5496"; band = "F2F5FB"; grey = "808080"; white = "FFFFFF"
thin = Side(style="thin", color="BFBFBF"); border = Border(thin, thin, thin, thin)

CASES = [
 ("SET-01","Setup","HR Admin","Sign in as HR Admin at http://localhost:5180. Check the top nav: Learning menu (Skill verifications, Skill inventory), Staffing/Workforce plans, the new Budgeting menu (Budget cycles, Department budgets, Payroll forecast, Variance, Settings) and the manager/team area.","All the listed menu entries are present."),
 # ---------- HCM_18 Phase A: skill verification (M419) ----------
 ("SKV-01","Skill verification (M419)","Employee","Sign in as an employee. Open Learning → Skill verifications → 'My requests' and submit a verification request for one of your competencies at a target level (attach evidence if the form offers it).","The request appears as PENDING in My requests."),
 ("SKV-02","Skill verification (M419)","Manager","Sign in as that employee's MANAGER. Open the pending verification queue and Approve the request.","Approval succeeds; the request becomes APPROVED and the employee's competency is upserted to the requested level (check the employee's competency list). A manager from a DIFFERENT hierarchy cannot approve it."),
 ("SKV-03","Skill verification (M419)","HR Admin","Submit another request as the employee, then as HR reject it with a reason.","Status becomes REJECTED; the employee's competency level is unchanged."),
 # ---------- HCM_18: skill inventory (M420) ----------
 ("SKI-01","Skill inventory (M420)","HR Admin","Learning → Skill inventory. Check the three tabs: inventory by department, critical-skill coverage, certification coverage.","Each tab populates: department rows show competency counts/avg levels; critical coverage reflects position required competencies; certification tab lists coverage + certificates expiring within 90 days."),
 ("SKI-02","Skill inventory scope (M420)","Manager","Sign in as a manager and open Skill inventory.","Only the manager's own team/department data is shown (hierarchy scope) — not the whole company."),
 # ---------- HCM_18: cert→skill link + types (M421) ----------
 ("CSK-01","Cert→skill link (M421)","HR Admin","Open a competency in the admin catalog — set its Skill type (e.g. TECHNICAL). Then edit an employee certification and link it to that competency. Mark/verify the certification as VERIFIED.","The competency shows its skill-type tag. On verification the employee is auto-awarded the linked competency at level 3 (visible in their competency list)."),
 # ---------- HCM_19 Phase B: hiring plan (M422) ----------
 ("HPL-01","Hiring plan (M422)","HR Admin","Open an APPROVED workforce plan (Staffing → Workforce plans) that has NEW_HIRE lines → 'Hiring plan' tab → 'Generate from plan'.","One hiring-plan line per NEW_HIRE plan line appears with status PLANNED and the target start date. Re-generating does not duplicate lines."),
 ("HPL-02","Hiring plan link (M422)","HR Admin","On a PLANNED hiring line use the link-vacancy action to attach an existing open vacancy.","The line stores the vacancy and flips to VACANCY_OPEN."),
 # ---------- HCM_19: attrition forecast (M423) ----------
 ("ATR-01","Attrition forecast (M423)","HR Admin","On the workforce plan open the 'Attrition forecast' tab and click Recalculate.","Rows appear per org-unit with expected exits and basis tags (HISTORICAL from last-12-months exits; CONTRACT_EXPIRY from contracts ending in the horizon). Recalculating replaces the previous rows (no duplicates)."),
 # ---------- HCM_19: plan→budget (M424) ----------
 ("PTB-01","Plan→budget transfer (M424)","HR Admin","On an APPROVED workforce plan click 'Transfer to budget' and confirm. Then check Compensation budgets for a new row referencing the plan. Click transfer again.","A compensation budget row is created from the plan totals (ref PLAN:<id>). The second attempt is rejected/disabled — transfer is idempotent."),
 # ---------- HCM_20 Phase C: budget cycles + department budgets (M425) ----------
 ("BCY-01","Budget cycles (M425)","HR Admin","Budgeting → Budget cycles → create cycle 'FY2027' (ANNUAL, next year's period). Set it OPEN.","The cycle appears with status OPEN. Editing a LOCKED/CLOSED cycle is rejected."),
 ("DBU-01","Department budgets (M425)","HR Admin","Budgeting → Department budgets → pick the OPEN cycle → add a budget for one department: salary 500000, training 20000, overtime 30000. Save, then edit and re-save.","The budget row saves (one per cycle+department — a duplicate for the same department is rejected/updates the same row); totals compute."),
 ("DBU-02","Budget approval (M425)","HR Admin","Submit the department budget for approval, then approve it from the Approvals inbox.","Submit → SUBMITTED and a BUDGET_APPROVAL workflow starts; approving flips it APPROVED (audited). Rejecting another budget returns it to DRAFT."),
 ("SEC-C1","Budget confidentiality","Employee / Manager","Sign in as a plain employee and as a manager. Try /budgets/cycles and /budgets/departments directly.","Denied/hidden for both — budget administration is HR/Finance-only."),
 # ---------- HCM_20: forecast (M426) ----------
 ("FCT-01","Payroll forecast (M426)","HR Admin","Budgeting → Payroll forecast. Set 12 months, growth 5%. ","A monthly line chart renders: base = current active payroll cost, rising with planned hires (hiring plan), falling with expected exits (attrition forecast), plus growth. Assumptions are listed."),
 # ---------- HCM_20: variance (M427) ----------
 ("VAR-01","Variance (M427)","HR Admin","Budgeting → Variance → pick a cycle that overlaps existing payroll runs.","Per-department rows show salary budget vs actual payroll cost with variance % and colour status (UNDER <90 green, WARNING 90–100 orange, OVER >100 red)."),
 ("VAR-02","Variance manager scope (M427)","Manager","Sign in as a department manager and open Variance.","The manager sees ONLY their own department's aggregate row — no other departments, no per-employee salary detail."),
 # ---------- HCM_20: control rules (M428) ----------
 ("BCR-01","Control rules (M428)","HR Admin","Budgeting → Settings → create a rule: trigger SALARY_CHANGE, action WARN, threshold 100%. Then submit a salary change that pushes the department over budget.","The change still submits, but a budget warning is recorded (audit/response)."),
 ("BCR-02","Control rules BLOCK (M428)","HR Admin","Change the rule action to BLOCK and submit another over-budget salary change (or an offer for trigger NEW_HIRE).","The operation is REJECTED with a clear budget-control message. Setting the rule inactive lets it pass again."),
 # ---------- HCM_21 Phase D: HR helpdesk (M429) ----------
 ("HDK-01","HR helpdesk (M429)","Employee","My Workspace → HR requests → submit a request: category PAYROLL_INQUIRY, priority NORMAL, subject + description.","The request appears in 'My requests' as OPEN with an SR-number and an SLA due date ~2 business days out (HIGH=1, LOW=5)."),
 ("HDK-02","HR helpdesk queue (M429)","HR Admin","Open the HR service queue, assign the request, start it, then resolve it with resolution notes.","Status flows OPEN→IN_PROGRESS→RESOLVED; the employee gets a notification; the queue can filter by status/category/overdue."),
 ("HDK-03","Grievance confidentiality (M429)","HR Specialist","Have an employee submit a GRIEVANCE request. Sign in as HR Specialist and look for it in the queue; then as HR Admin.","The grievance is hidden from HR Specialist — visible to HR Admin only."),
 # ---------- HCM_21: announcements (M430) ----------
 ("ANN-01","Announcements (M430)","HR Admin","Create an announcement (audience ALL, active, publish window covering today) and a second one targeted at a department the test employee is NOT in.","Both save in the admin list."),
 ("ANN-02","Announcements audience (M430)","Employee","Sign in as the employee and open My Workspace.","The ALL announcement shows on the dashboard; the other-department one does NOT."),
 # ---------- HCM_21: team calendar ESS (M431) ----------
 ("TCA-01","Team calendar ESS (M431)","Manager / Employee","Sign in as a manager → My Workspace → Team calendar widget/tab. Then sign in as a non-manager employee.","The manager sees their team's leave calendar; the non-manager doesn't get the widget (or an empty/'not a manager' state) — never another team's data."),
 # ---------- HCM_22 Phase E: movement requests (M432) ----------
 ("MOV-01","Movement request (M432)","Manager","Open Manager → Movements → create a PROMOTION request for one of YOUR reports (proposed position/grade, effective date, justification). Submit.","The MV-numbered request goes SUBMITTED and an EMP_MOVEMENT approval starts. Creating one for an employee outside your team is rejected."),
 ("MOV-02","Movement approve (M432)","HR Admin","Approve the movement request from the Approvals inbox.","Status becomes APPROVED (audited) and HR is notified to execute the change via the normal transfer/contract tooling (the request does NOT auto-execute the move)."),
 ("MOV-03","Movement salary privacy (M432)","Manager","As the requesting manager, view the movement request that has a proposed salary.","The proposed salary is NOT visible to the manager (HR-only field); HR Admin sees it."),
 # ---------- HCM_22: comp visibility toggle (M433) ----------
 ("CVT-01","Comp visibility toggle (M433)","HR Admin / Manager","With the tenant setting manager_can_view_salary OFF (default), sign in as a manager and look for team compensation. Then as HR Admin turn the setting ON and re-check.","OFF → no compensation column/section for the manager (API denies). ON → the manager sees salary for OWN reports only."),
 # ---------- HCM_22: manager analytics (M434) ----------
 ("MAN-01","Manager analytics (M434)","Manager","Open Manager → Analytics.","Stat cards/charts show the manager's OWN team only: headcount, 12-month turnover, absence rate, overtime hours, training completion. Numbers are plausible vs the team's data."),
 # ---------- HCM_22: inbox enhancements (M435) ----------
 ("INB-01","Inbox filters + SLA (M435)","Manager","Open the Approvals inbox. Use the new type filter (e.g. Leave only) and look for SLA badges on overdue items.","Filtering narrows the list by workflow type; overdue items carry a red SLA badge."),
 ("INB-02","Bulk approve (M435)","Manager","Select 2+ pending items with checkboxes and click 'Approve selected'.","A confirm modal runs the action per item and reports per-item results; items you are not allowed to act on FAIL individually (permission re-checked per item) without blocking the rest."),
 # ---------- Cross-cutting ----------
 ("SEC-X1","ESS scoping","Employee","As a plain employee try: /budgets/variance, /manager/movements, /manager/analytics, the HR service queue URL, and another employee's HR request by id.","All denied or empty — employees reach only their own ESS surfaces."),
 ("SEC-X2","Tenant/hierarchy spot-check","Manager","As a manager, hit the skill-inventory and variance APIs with crafted params (another department's org-unit id).","Responses stay limited to the manager's own scope — no cross-hierarchy data."),
]

wb = Workbook()

# Instructions
ws = wb.active; ws.title = "Instructions"; ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 26; ws.column_dimensions["C"].width = 100
def put(r,l,v,bold=True):
    b=ws.cell(r,2,l); b.font=Font(name=FONT,bold=bold,size=11,color=navy if bold else "000000"); b.alignment=Alignment(vertical="top",wrap_text=True)
    c=ws.cell(r,3,v); c.font=Font(name=FONT,size=11); c.alignment=Alignment(vertical="top",wrap_text=True)
ws.cell(1,2,"Skills / Workforce / Budgeting / ESS / MSS (HCM_18-22) — UAT Script").font=Font(name=FONT,bold=True,size=16,color=navy)
ws.cell(2,2,"HCM_18-22 bundle  •  front-end (browser) testing  •  grows as each phase ships").font=Font(name=FONT,size=11,color=grey)
for r,l,v in [
 (4,"How to use","Open the 'Test Cases' sheet. Do exactly what the Steps say, compare to the Expected Result, and pick Pass / Fail / Blocked / Not Run in the Result column. Add anything unusual under Tester Notes. Fill the Sign-off sheet at the end."),
 (6,"Application URL","http://localhost:5180 — sign in first. Features live under the Learning, Staffing, Budgeting, My Workspace and Manager menus."),
 (7,"Delivered","HCM_18: skill verification workflow (M419), skill inventory reports (M420), cert→skill auto-link + skill types (M421). HCM_19: hiring plan from workforce plan (M422), attrition forecast (M423), plan→budget transfer (M424). HCM_20: budget cycles + department budgets + approval (M425), payroll forecast (M426), budget-vs-actual variance (M427), budget control rules WARN/BLOCK (M428). HCM_21: HR helpdesk with SLA (M429), announcements (M430), team calendar in ESS (M431). HCM_22: transfer/promotion movement requests (M432), team comp visibility toggle (M433), manager analytics (M434), inbox filters + SLA badges + bulk approve (M435)."),
 (9,"Logins you will need","HR Admin, HR Specialist (grievance-confidentiality check), Manager (hierarchy + MSS checks), Employee (ESS + scoping checks). If you only have an admin login, mark role-restriction rows Blocked with a note."),
 (11,"Result values","Pass = worked as expected.  Fail = did not match (add a note).  Blocked = could not run (missing login/data).  Not Run = skipped."),
 (12,"CONFIDENTIALITY rules under test","(1) Budget data is HR/Finance-only; a manager sees at most their OWN department's variance aggregate. (2) Managers cannot see team salaries unless HR enables the tenant toggle — and then only for their own reports. (3) Movement-request proposed salary is HR-only. (4) Grievance helpdesk requests are HR-Admin-only. Any leak = automatic Fail + report immediately."),
]:
    put(r,l,v)
ws.row_dimensions[4].height=45; ws.row_dimensions[7].height=90; ws.row_dimensions[12].height=60

# Test Cases
tc = wb.create_sheet("Test Cases"); tc.sheet_view.showGridLines=False
headers=["Test ID","Feature Area","Login As","Test Steps (what to click / type)","Expected Result","Result","Tester Notes"]
widths=[12,26,18,62,58,12,30]
for i,(h,w) in enumerate(zip(headers,widths),1):
    c=tc.cell(1,i,h); c.font=Font(name=FONT,bold=True,color=white,size=11); c.fill=PatternFill("solid",fgColor=navy)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
    tc.column_dimensions[get_column_letter(i)].width=w
tc.freeze_panes="A2"; tc.row_dimensions[1].height=26
for idx,(tid,feat,role,steps,exp) in enumerate(CASES):
    r=idx+2; banded=(idx%2==1)
    for col,v in enumerate([tid,feat,role,steps,exp,"",""],1):
        c=tc.cell(r,col,v)
        c.font=Font(name=FONT,size=11,bold=(col==1),color=(navy if col==1 else "000000"))
        c.alignment=Alignment(vertical="top",wrap_text=True,horizontal=("center" if col in(1,6) else "left"))
        c.border=border
        if banded and col!=6: c.fill=PatternFill("solid",fgColor=band)
    tc.cell(r,6).fill=PatternFill("solid",fgColor="FFF7E6")
    longest=max(len(steps),len(exp)); tc.row_dimensions[r].height=max(30,min(160,(longest//60+1)*15+12))
dv=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True)
dv.add(f"F2:F{len(CASES)+1}"); tc.add_data_validation(dv)

# Sign-off
so=wb.create_sheet("Sign-off"); so.sheet_view.showGridLines=False
for col,w in zip("ABCDE",[3,44,16,22,18]): so.column_dimensions[col].width=w
so.cell(1,2,"HCM_18-22 bundle UAT — Sign-off").font=Font(name=FONT,bold=True,size=15,color=navy)
for i,h in enumerate(["Feature Area","Result","Tester","Date"],2):
    c=so.cell(3,i,h); c.font=Font(name=FONT,bold=True,color=white); c.fill=PatternFill("solid",fgColor=blue); c.border=border
areas=["Skill verification (SKV)","Skill inventory (SKI)","Cert→skill link + types (CSK)",
       "Hiring plan (HPL)","Attrition forecast (ATR)","Plan→budget transfer (PTB)",
       "Budget cycles + department budgets (BCY/DBU)","Payroll forecast (FCT)",
       "Budget variance (VAR)","Budget control rules (BCR)",
       "HR helpdesk (HDK)","Announcements (ANN)","Team calendar ESS (TCA)",
       "Movement requests (MOV)","Comp visibility toggle (CVT)","Manager analytics (MAN)",
       "Inbox enhancements (INB)","Permissions + confidentiality (SEC)"]
dv2=DataValidation(type="list",formula1='"Pass,Fail,Blocked,Not Run"',allow_blank=True); so.add_data_validation(dv2)
for j,a in enumerate(areas):
    r=4+j; so.cell(r,2,a).font=Font(name=FONT,size=11)
    for col in range(2,6): so.cell(r,col).border=border
    so.cell(r,3).fill=PatternFill("solid",fgColor="FFF7E6"); dv2.add(so.cell(r,3))
fr=4+len(areas)+2
so.cell(fr,2,"Overall decision (SHIP / DON'T SHIP):").font=Font(name=FONT,bold=True,size=12,color=navy)
so.cell(fr,3).fill=PatternFill("solid",fgColor="FFF7E6"); so.cell(fr,3).border=border
so.cell(fr+2,2,"Tester name:").font=Font(name=FONT,size=11); so.cell(fr+3,2,"Date:").font=Font(name=FONT,size=11)

wb.save(OUT)
print("WROTE",OUT,"cases:",len(CASES))
